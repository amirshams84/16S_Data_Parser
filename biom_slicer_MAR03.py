# ################################### INFO ##################################### #
# 16S BiomeSlicer 1.0
# Author: Amir Shams
# Email: amir.shams84@gmail.com
# ################################### IMPORTED LIBRARY ######################### #


import sys
import os
import errno
import datetime
import signal
import logging as log
import time
import subprocess
import traceback
import itertools
import argparse
import multiprocessing
import platform
import pandas
import numpy
import csv
import shutil
import decimal
import zipfile
import random
import collections
import operator
import plotly
import plotly.graph_objs as PLOTLY_GO
import biom
import openpyxl
import xlsxwriter
import openpyxl.chart as OPEN_CH
import openpyxl.utils.dataframe as OPEN_DF
import openpyxl.chart.data_source as OPEN_DS
import openpyxl.chart.series_factory as OPEN_SERIES
# ################################### GLOBALS ################################## #


CHECK_MARK = "OK"
FAILED_MARK = ":("
DEFAULT_OUTPUTDIR = "/BIOM_SLICER_OUTPUTDIR/"
DEFAULT_TESTDIR = "/BIOM_SLICER_TESTDIR/"
DEFAULT_EXECDIR = "/BIOM_SLICER_EXECDIR/"
DEFAULT_PROCESSORS = str(multiprocessing.cpu_count())
DEFAULT_PREFIX = "BIOM_SLICER"
NO_BETA = True
CURRENT_PATH = "./"
TAXONOMY_EXIST = False
global DESIGN_EXIST
# ################################### OBJECTS ################################## #


class mothur_process:
	def __init__(self, mothur_input_dictionary):
		for var_name, var_value in mothur_input_dictionary.items():
			setattr(self, var_name, var_value)

	def build_mothur_command(self):
		space = " "
		string = ''
		if hasattr(self, 'nohup_in'):
			string += self.nohup_in + space
		string += self.mothur_exec_path + space + '"#set.dir(output=' + self.outputdir + ');' + space
		string += 'set.logfile(name=mothur.' + self.command + '.logfile);' + space
		string += 'set.current(processors=' + self.processors + ');' + space
		string += self.command + '('
		if hasattr(self, 'parameters') and len(self.parameters) > 0:
			for each_element in self.parameters:
				string += each_element
		string += ');"'
		if hasattr(self, 'nohup_out'):
			string += space + self.nohup_out
		if hasattr(self, 'pid_file'):
			string += ' echo $! > ' + self.pid_file
		report(string)
		print string
		return string

	def execute_mothur_command(self):
		exec_dict = {}
		exec_dict = self.execute([self.build_mothur_command()])
		if exec_dict['exitCode'] != 0:
			print "ERROR occurred!!!!"
			return (False, exec_dict)
		else:
			return(True, exec_dict)

	def execute(self, command, ** kwargs):
		assert isinstance(command, list), "Expected 'command' parameter to be a list containing the process/arguments to execute. Got %s of type %s instead" % (command, type(command))
		assert len(command) > 0, "Received empty list of parameters"
		retval = {
					"exitCode": -1,
					"stderr": u"",
					"stdout": u"",
					"execTime": datetime.timedelta(0),
					"command": None,
					"pid": None
				}
		retval["command"] = command
		log.info("::singleProcessExecuter > At %s, executing \"%s\"" % (datetime.datetime.now(), " ".join(command)))
		cwd = kwargs.get("cwd", os.getcwd())
		#user = kwargs.get("user", os.getuid)
		sheel = kwargs.get("shell", True)
		startDatetime = datetime.datetime.now()
		myPopy = subprocess.Popen(command, cwd=cwd, preexec_fn=os.seteuid(os.getuid()), shell=sheel, stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE)
		retval["pid"] = myPopy.pid
		log.debug("::singleProcessExecuter > Command \"%s\" got pid %s" % (" ".join(command), myPopy.pid))
		try:
			retval["stdout"], retval["stderr"] = myPopy.communicate()
			myPopy.wait()
		except OSError, osErr:
			log.debug("::singleProcessExecuter > Got %s %s in myPopy.communicate() when trying get output of command %s. It is probably a bug (more info: http://bugs.python.org/issue1731717)" % (osErr, type(osErr), command[0]))
		except Exception, e:
			log.warn("::singleProcessExecuter > Got %s %s when trying to get stdout/stderr outputs of %s" % (type(e), e, " ".join(command)))
			log.debug("::singleProcessExecuter > Got %s %s when trying to get stdout/stderr outputs of %s. Showing traceback:\n%s" % (type(e), e, " ".join(command), traceback.format_exc()))
			raise
		retval["exitCode"] = myPopy.returncode
		retval["execTime"] = datetime.datetime.now() - startDatetime
		return retval
# ################################### EXECUTIONS ############################### #


def execute_functions(function_name, processors, outputdir, thread_type, func_mode, *args):
	print "Execution started at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution started at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	list_of_pid = []
	list_of_stderr = []
	list_of_stdout = []
	if thread_type == 'fork':
		threads = int(processors) + 1
		processors = '1'
	elif thread_type == 'multi':
		threads = 2
	for thread in range(1, threads):
		pid_file = 'run.pid' + str(thread)
		stderr_file = 'stderr' + str(thread)
		stdout_file = 'stdout' + str(thread)
		run_pid = outputdir + pid_file
		stderr = outputdir + stderr_file
		stdout = outputdir + stdout_file
		flag = function_name(processors, outputdir, stderr, stdout, run_pid, *args)
		if flag is False:
			sys.exit(2)
		list_of_pid.append(pid_file)
		list_of_stderr.append(stderr_file)
		list_of_stdout.append(stdout_file)
	flag, stderr = process_monitor(list_of_pid, list_of_stderr, list_of_stdout, outputdir, threads, func_mode)
	if flag is False:
		print "Process monitor failed."
	else:
		print "Execution completed at ", time.strftime("%Y-%m-%d %H:%M:%S")
		report("Execution completed at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	return (True, stderr)


def process_monitor(pid_list, stderr_list, stdout_list, outputdir, threads, mode):

	finished_flag = False
	flag_list = {}
	for each_pid in pid_list:
		flag_list[each_pid] = False
	toolbar_width = threads
	sys.stdout.write("[%s]" % (" " * toolbar_width))
	sys.stdout.flush()
	sys.stdout.write("\b" * (toolbar_width + 1))  # return to start of line, after '['
	while finished_flag is False:
		for pid_file, stderr_file, stdout_file in itertools.izip(pid_list, stderr_list, stdout_list):
			f = open(outputdir + pid_file)
			the_pid = int(f.read().rstrip())
			if pid_exists(the_pid) is False:
				sys.stdout.write("OK")
				sys.stdout.flush()
				flag_list[pid_file] = True
				flag, stderr = validate_execution(stderr_file, stdout_file, outputdir, mode)
				if flag is False:
					sys.stdout.write(":(")
					report("[:()]")
					print "Error in result of this thread: ", str(the_pid)
					error("Error in result of this thread: " + str(the_pid))
					print "All generated threads killed."
					error("All generated threads killed.")
					kill_pid_list(pid_list, outputdir)
					#print stderr
					print "ABORTING!!!"
					report("ABORTING!!!")
					sys.exit(2)
			if False in flag_list.values():
				finished_flag = False
			else:
				finished_flag = True
		time.sleep(1)
	sys.stdout.write("\n")
	report("[OK]")
	return (True, stderr)


def pid_exists(pid):
	if pid < 0:
		return False
	if pid == 0:
		# According to "man 2 kill" PID 0 refers to every process
		# in the process group of the calling process.
		# On certain systems 0 is a valid PID but we have no way
		# to know that in a portable fashion.
		raise ValueError('invalid PID 0')
	try:
		os.kill(pid, 0)
	except OSError as err:
		if err.errno == errno.ESRCH:
			# ESRCH == No such process
			return False
		elif err.errno == errno.EPERM:
			# EPERM clearly means there's a process to deny access to
			return True
		else:
			# According to "man 2 kill" possible error values are
			# (EINVAL, EPERM, ESRCH)
			raise
	else:
		return True


def validate_execution(stderr, stdout, outputdir, mode):
	if mode == 'usearch':
		f = open(outputdir + stdout, 'rU')
		data = f.read()
		#print data
		if '---Fatal error---' in data or 'Invalid command line' in data:
			return (False, data)
		else:
			return(True, data)
	elif mode == 'mothur':
		f = open(outputdir + stdout, 'rU')
		data = f.read()
		if '[ERROR]:' in data:
			return (False, data)
		else:
			return(True, data)
	elif mode == 'mafft':
		f = open(outputdir + stdout, 'rU')
		data = f.read()
		if 'No such file or directory' in data or 'cannot open file' in data:
			return (False, data)
		else:
			return(True, data)


def kill_pid_list(pid_list, outputdir):
	for pid in pid_list:
		f = open(outputdir + pid)
		the_pid = int(f.read().rstrip())
		try:
			os.kill(the_pid, signal.SIGTERM)
		except OSError:
			pass
		print the_pid, "Killed!"
	return True


def execute(command, ** kwargs):
	assert isinstance(command, list), "Expected 'command' parameter to be a list containing the process/arguments to execute. Got %s of type %s instead" % (command, type(command))
	assert len(command) > 0, "Received empty list of parameters"
	retval = {
			"exitCode": -1,
			"stderr": u"",
			"stdout": u"",
			"execTime": datetime.timedelta(0),
			"command": None,
			"pid": None
		}
	retval["command"] = command
	log.info("::singleProcessExecuter > At %s, executing \"%s\"" % (datetime.datetime.now(), " ".join(command)))
	# print("::singleProcessExecuter > At %s, executing \"%s\"" % (datetime.datetime.now(), " ".join(parameter)))
	cwd = kwargs.get("cwd", os.getcwd())
	sheel = kwargs.get("shell", True)
	startDatetime = datetime.datetime.now()
	myPopy = subprocess.Popen(command, cwd=cwd, preexec_fn=os.seteuid(os.getuid()), shell=sheel, stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE)
	retval["pid"] = myPopy.pid
	log.debug("::singleProcessExecuter > Command \"%s\" got pid %s" % (" ".join(command), myPopy.pid))
	try:
		retval["stdout"], retval["stderr"] = myPopy.communicate()
		myPopy.wait()
	except OSError, osErr:
		log.debug("::singleProcessExecuter > Got %s %s in myPopy.communicate() when trying get output of command %s. It is probably a bug (more info: http://bugs.python.org/issue1731717)" % (osErr, type(osErr), command[0]))
	except Exception, e:
		log.warn("::singleProcessExecuter > Got %s %s when trying to get stdout/stderr outputs of %s" % (type(e), e, " ".join(command)))
		log.debug("::singleProcessExecuter > Got %s %s when trying to get stdout/stderr outputs of %s. Showing traceback:\n%s" % (type(e), e, " ".join(command), traceback.format_exc()))
		raise
	retval["exitCode"] = myPopy.returncode
	retval["execTime"] = datetime.datetime.now() - startDatetime
	return retval
# ################################### LOGGING & REPORTING ##################### #


def error(report_string):
	f = open(report_file, "a")
	f.write('###############################  ERROR   ###################################\n')
	f.write(report_string)
	f.write("\n")
	f.write('############################################################################\n')
	f.close()


def report(report_string):
	f = open(report_file, "a")
	f.write(report_string.encode('utf8'))
	f.write("\n")
	f.close()
# ################################### UTILITIES ############################## #


def check_it_and_remove_it(filename, noreport=False):
	try:
		os.remove(filename)
		if noreport is False:
			pass
	except OSError:
		pass


def isFileExist(fname):
	# check the exitence/access/path of a file
	if os.path.isfile(fname) and os.path.exists(fname) and os.access(fname, os.R_OK):
		return True
	else:
		return False


def list_to_string(query_list, delimiter=None):
	# convert list to string by specified delimiter
	if delimiter is None:
		delimiter = '\t'
	return delimiter.join(map(str, query_list))


def slugify(target_word):
	text = target_word.replace(' ', r'_').replace('\\', r'_').replace('`', r'_').replace('{', r'_').replace('}', r'_').replace('[', r'_').replace(']', r'_').replace('(', r'_').replace(')', r'_').replace('>', r'_').replace('#', r'_').replace('+', r'_').replace('-', r'_').replace('.', r'_').replace('!', r'_').replace('$', r'_').replace("'", r'_').replace('"', r'_').replace('\/', r'_').replace('__', r'_').replace('__', r'_').replace('__', r'_').replace('__', r'_').replace('__', r'_').replace('__', r'_').replace('__', r'_')
	return text


def get_pandas_DATAFRAME(file_PATH):
	extension = get_extension(file_PATH)
	if extension in ['txt', 'tsv', 'csv']:
		return pandas.read_table(file_PATH, low_memory=False, encoding='utf-8', skip_blank_lines=True, error_bad_lines=False)
	elif extension in ['xlsx', 'xls']:
		return pandas.read_excel(file_PATH, sheetname=None)
	else:
		print "Unknow extension"
		error("Unknow extension")
		print "ABORTING!!!"
		error("ABORTING!!!")
		sys.exit(2)


def get_extension(file_PATH):
	#Get Extension
	return file_PATH.split('.')[-1].lower()


def match_two_list(list_A, list_B):
	#Match two list
	return set(list_A) & set(list_B)


def round_float(value):
	if isfloat(value) is False:
		return value
	else:
		return round(decimal.Decimal(value), 3)


def percentage(part, whole):
	#return percentage
	return 100 * float(part) / float(whole)


def isfloat(x):
	try:
		float(x)
	except ValueError:
		return False
	else:
		return True


def remove_extension_files(outputdir, extension):
	extension_list = []
	extension_list.append(extension)
	scanned_container = []
	flag = scandirs(outputdir, scanned_container, extension_list, 'partial')
	print "Scanning to find", extension, "Files.."
	if flag is False:
		print "Failed :("
		print "This extension is not available: ", extension_list
		sys.exit(2)
	else:
		print "NICE :) Found them"
		counter = 1
		for file in scanned_container:
			#print "Removing File#", str(counter), ":", file
			counter += 1
			check_it_and_remove_it(file)
	return True


def scandirs(path, container, ext_list, mode=None):
	# scan a spath and grab all files by specified extension
	for root, dirs, names in os.walk(path):
		for currentFile in names:
			path, absname, ext = split_file_name(currentFile)
			if mode is None:
			# Looking for exactly the same extension
				if ext in ext_list:
					container.append(os.path.join(root, currentFile))
			elif mode == 'multiple':
				for each_ext in ext_list:
					if ext in each_ext:
						container.append(os.path.join(root, currentFile))
			elif mode == 'partial':
				# when this extension is part of the actual extension of files
				for each_ext in ext_list:
					if each_ext in ext:
						container.append(os.path.join(root, currentFile))
	if len(container) < 1:
		return False
	return True


def split_file_name(file):
	path = os.path.dirname(file) + '/'
	name = os.path.basename(file)
	if '.' in name:
		ext = '.' + '.'.join(name.split('.')[1:])
		absname = name.split('.')[0]
	else:
		ext = 'no_extension'
		absname = name
	return (path, absname, ext)


def write_string_down(new_string, file_name):
	f = open(file_name, 'w')
	f.write(new_string)
	f.close()
	return True


def zip_it(target_file_NAME, zip_file_NAME):
	zip_file_HANDLE = zipfile.ZipFile(zip_file_NAME, 'w')
	zip_file_HANDLE.write(target_file_NAME)
	zip_file_HANDLE.close()
	return True


def add_extension(file_PATH, extension):
	#Add Extension
	os.rename(file_PATH, file_PATH + extension)
	return file_PATH + extension


def numpy_percentile(freq_list, threshold):
	numpy_array = numpy.array(freq_list)
	return numpy.percentile(numpy_array, threshold)


def filter_list(any_LIST, Element):
	filtered_LIST = []
	for each_element in any_LIST:
		if each_element == Element:
			continue
		else:
			filtered_LIST.append(each_element)
	return filtered_LIST
# ################################### CONVERTER ############################## #


def biom_to_excel_converter(Biom_File_PATH, Design_File_PATH, Excel_File_PATH):
	#Step1: Get biom handle
	Biom_HANDLE = biom.load_table(Biom_File_PATH)
	Excel_writer = pandas.ExcelWriter(Excel_File_PATH, engine='xlsxwriter')
	# ######################################################################
	# Biome_Data
	# ######################################################################
	
	Biom_DICT = {}
	Biom_LIST = []
	Sample_ID_LIST = Biom_HANDLE.ids(axis='sample')
	Biom_ID_LIST = Biom_HANDLE.ids(axis='observation')
	Biom_DICT['OTU_ID'] = []
	Biom_DICT['Biom_ID'] = []
	Biom_LIST = ['OTU_ID']
	Biom_LIST.append('Biom_ID')
	if Biom_HANDLE.metadata(axis='observation') is not None:
		metadata_OTU_OBJECT = Biom_HANDLE.metadata(axis='observation')
		for each_extra_headers in metadata_OTU_OBJECT[0].keys():
			Biom_DICT[each_extra_headers] = []
			Biom_LIST.append(each_extra_headers)
	Biom_LIST.extend(Sample_ID_LIST)
	for each_Sample in Sample_ID_LIST:
		Biom_DICT[each_Sample] = Biom_HANDLE.data(each_Sample, axis='sample').tolist()
	
	each_OTU_metadata_DICT = {}
	OTU_ID_counter = 1
	for each_Biom_ID in Biom_ID_LIST:
		OTU_ID = 'OTU_ID_' + str(OTU_ID_counter).zfill(6)
		Biom_DICT['OTU_ID'].append(OTU_ID)
		if Biom_HANDLE.metadata(each_Biom_ID, axis='observation') is not None:
			each_Biom_ID_metadata_DICT = dict(Biom_HANDLE.metadata(each_Biom_ID, axis='observation'))
			metadata_keys = each_Biom_ID_metadata_DICT.keys()
			for each_metadata_key in metadata_keys:
				if type(each_Biom_ID_metadata_DICT[each_metadata_key]) is list:
					Biom_DICT[each_metadata_key].append(list_to_string(each_Biom_ID_metadata_DICT[each_metadata_key], ';'))
				else:
					Biom_DICT[each_metadata_key].append(each_Biom_ID_metadata_DICT[each_metadata_key])
		Biom_DICT['Biom_ID'].append(each_Biom_ID)
		OTU_ID_counter += 1
	Biom_MATRIX = {}
	Biom_MATRIX_LIST = []
	for each_Sample in Sample_ID_LIST:
		Biom_MATRIX[each_Sample] = dict(zip(Biom_DICT['OTU_ID'], Biom_DICT[each_Sample]))
		Biom_MATRIX_LIST.append(each_Sample)
	#print Biom_MATRIX
	#{'PC.355': {'OTU_ID_000343': 1.0, 'OTU_ID_000342': 0.0, 'OTU_ID_000101': 0.0,
	#Pandas_DataFrame = pandas.DataFrame.from_dict(Biom_DICT, orient='columns', dtype='float')
	Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(Biom_DICT, Biom_LIST, the_INDEX='OTU_ID')
	Pandas_DataFrame.to_excel(Excel_writer, sheet_name='Biome_Data', columns=pandas_DataFrame_LIST, header=True, index=True)
	print "BIOME_Data created!!!"
	# ######################################################################
	# Sample_metadata
	# ######################################################################
	#Step4: Check for sample metadata
	if Biom_HANDLE.metadata(axis='sample') is not None:
		metadata_sample_OBJECT = Biom_HANDLE.metadata(axis='sample')
		Sample_DICT = {}
		Sample_LIST = ['Sample_ID', 'Sample_Name']
		Sample_DICT['Sample_ID'] = []
		Sample_DICT['Sample_Name'] = []
		for each_extra_headers in metadata_sample_OBJECT[0].keys():
			Sample_DICT[each_extra_headers] = []
			Sample_LIST.append(each_extra_headers)
		each_Sample_metadata_DICT = {}
		Sample_ID_counter = 1
		for each_Sample in Sample_ID_LIST:
			Sample_DICT['Sample_ID'].append('Sample_ID_' + str(Sample_ID_counter).zfill(3))
			Sample_DICT['Sample_Name'].append(each_Sample)
			each_Sample_metadata_DICT = dict(Biom_HANDLE.metadata(each_Sample, axis='sample'))
			metadata_keys = each_Sample_metadata_DICT.keys()
			for each_metadata_key in metadata_keys:
				if type(each_Sample_metadata_DICT[each_metadata_key]) is list:
					Sample_DICT[each_metadata_key].append(list_to_string(each_Sample_metadata_DICT[each_metadata_key], ';'))
				else:
					Sample_DICT[each_metadata_key].append(each_Sample_metadata_DICT[each_metadata_key])
			Sample_ID_counter += 1
		Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(Sample_DICT, Sample_LIST, the_INDEX='Sample_ID')
		Pandas_DataFrame.to_excel(Excel_writer, sheet_name='Sample_metadata', columns=pandas_DataFrame_LIST, header=True, index=True)
	else:
		Sample_DICT = {}
		Sample_LIST = ['Sample_ID', 'Sample_Name', 'Label']
		Sample_DICT['Sample_ID'] = []
		Sample_DICT['Sample_Name'] = []
		Sample_DICT['Label'] = []
		Sample_ID_counter = 1
		for each_Sample in Sample_ID_LIST:
			Sample_DICT['Sample_ID'].append('Sample_ID_' + str(Sample_ID_counter).zfill(3))
			Sample_DICT['Sample_Name'].append(each_Sample)
			Sample_DICT['Label'].append('BiomeSlicer')
			Sample_ID_counter += 1
		Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(Sample_DICT, Sample_LIST, the_INDEX='Sample_ID')
		Pandas_DataFrame.to_excel(Excel_writer, sheet_name='Sample_metadata', columns=pandas_DataFrame_LIST, header=True, index=True)
	print "Sample_metadata created!!!"
	# ######################################################################
	# OTU_metadata
	# ######################################################################
	#Step5: Check for OTU metadata
	if Biom_HANDLE.metadata(axis='observation') is not None:
		metadata_OTU_OBJECT = Biom_HANDLE.metadata(axis='observation')
		OTU_DICT = {}
		OTU_LIST = []
		#For taxonomy infere usage
		OTU_ID_LIST = []
		OTU_ID_DICT = {}
		OTU_DICT['OTU_ID'] = []
		OTU_LIST.append('OTU_ID')
		for each_extra_headers in metadata_OTU_OBJECT[0].keys():
			OTU_DICT[each_extra_headers] = []
			OTU_LIST.append(each_extra_headers)
		each_OTU_metadata_DICT = {}
		for each_OTU_ID, each_Biom_ID in itertools.izip(Biom_DICT['OTU_ID'], Biom_DICT['Biom_ID']):
			OTU_DICT['OTU_ID'].append(each_OTU_ID)
			each_OTU_metadata_DICT = dict(Biom_HANDLE.metadata(each_Biom_ID, axis='observation'))
			metadata_keys = each_OTU_metadata_DICT.keys()
			for each_metadata_key in metadata_keys:
				if each_metadata_key == 'taxonomy':
					corrected_taxonomy_string = remove_ambiguity_from_taxonomy(each_OTU_metadata_DICT[each_metadata_key])
					OTU_DICT[each_metadata_key].append(corrected_taxonomy_string)
					OTU_ID_DICT[each_OTU_ID] = corrected_taxonomy_string
					OTU_ID_LIST.append(each_OTU_ID)
				elif type(each_OTU_metadata_DICT[each_metadata_key]) is list:
					OTU_DICT[each_metadata_key].append(list_to_string(each_OTU_metadata_DICT[each_metadata_key], ';'))
				else:
					OTU_DICT[each_metadata_key].append(each_OTU_metadata_DICT[each_metadata_key])
		Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(OTU_DICT, OTU_LIST, the_INDEX='OTU_ID')
		Pandas_DataFrame.to_excel(Excel_writer, sheet_name='OTU_metadata', columns=pandas_DataFrame_LIST, header=True, index=True)
		OTU_MATRIX = {}
		OTU_MATRIX_LIST = []
		OTU_MATRIX = dict(zip(OTU_DICT['OTU_ID'], OTU_DICT['taxonomy']))
		OTU_MATRIX_LIST.extend(OTU_DICT['OTU_ID'])
	else:
		OTU_DICT = {}
		OTU_LIST = ['OTU_ID', 'taxonomy']
		OTU_DICT['OTU_ID'] = []
		OTU_DICT['taxonomy'] = []
		#For taxonomy infere usage
		OTU_ID_LIST = []
		OTU_ID_DICT = {}
		for each_OTU_ID, each_Biom_ID in itertools.izip(Biom_DICT['OTU_ID'], Biom_DICT['Biom_ID']):
			OTU_DICT['OTU_ID'].append(each_OTU_ID)
			OTU_DICT['taxonomy'].append(each_Biom_ID)
			OTU_ID_DICT[each_OTU_ID] = each_Biom_ID
			OTU_ID_LIST.append(each_OTU_ID)
		Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(OTU_DICT, OTU_LIST, the_INDEX='OTU_ID')
		Pandas_DataFrame.to_excel(Excel_writer, sheet_name='OTU_metadata', columns=pandas_DataFrame_LIST, header=True, index=True)
		OTU_MATRIX = {}
		OTU_MATRIX_LIST = []
		OTU_MATRIX = dict(zip(OTU_DICT['OTU_ID'], OTU_DICT['taxonomy']))
		OTU_MATRIX_LIST.extend(OTU_DICT['OTU_ID'])
	#print OTU_MATRIX
	#{'OTU_ID_000343': 'k_bacteria;p_firmicutes;c_clostridia;o_clostridiales;f_lachnospiraceae;*f_lachnospiraceae;*f_lachnospiraceae;',
	print "OTU_metadata created!!!"
	# ######################################################################
	# Design
	# ######################################################################
	if Design_File_PATH is not None:
		Design_DICT = {}
		Design_LIST = []
		Design_DICT, Design_LIST = any_file_to_vertical_dict_converter(Design_File_PATH)

		key_column = ''
		key_column_len = 0
		#print SAMPLE_metadata_DATAFrame_DICT['SAMPLE_ID']
		for each_Column in Design_LIST:
			match_rank = len(match_two_list(Sample_DICT['Sample_Name'], Design_DICT[each_Column]))
			if match_rank > key_column_len:
				key_column_len = match_rank
				key_column = each_Column
		if key_column_len > 1:
			print "Key Column for design file parsing is: ", key_column
			Category_DICT = {}
			Category_LIST = ['Sample_Name']
			
			Category_DICT['Sample_Name'] = []
			for each_Sample in Design_DICT[key_column]:
				Category_DICT['Sample_Name'].append(each_Sample)
			Design_LIST.remove(key_column)
			for each_Column in Design_LIST:
				Category_LIST.append(each_Column)
				Category_DICT[each_Column] = Design_DICT[each_Column]
			Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(Category_DICT, Category_LIST, the_INDEX=None)
			Pandas_DataFrame.to_excel(Excel_writer, sheet_name='Design', columns=pandas_DataFrame_LIST, header=True, index=False)
			print "Design created!!!"
			DESIGN_EXIST = True
		else:
			print "Design_file Does not match with data please check the Sample_ID column"
			DESIGN_EXIST = False
	else:
		DESIGN_EXIST = False
	# ######################################################################
	# TAX_metadata
	# ######################################################################
	TAX_DICT = {}
	TAX_DICT['TAX_ID'] = []
	TAX_DICT['TAX_LEVEL'] = []
	TAX_DICT['TAXONOMY'] = []
	TAX_LIST = []
	TAX_LIST = ['TAX_ID', 'TAX_LEVEL', 'TAXONOMY']
	TAX_ID_counter = 1
	for each_OTU in OTU_MATRIX_LIST:
		TAX_ID = 'TAX_ID_' + str(TAX_ID_counter).zfill(6)
		TAX_DICT['TAX_ID'].append(TAX_ID)
		TAX_DICT['TAX_LEVEL'].append('OTU')
		TAX_DICT['TAXONOMY'].append(OTU_MATRIX[each_OTU])
		TAX_ID_counter += 1
	###############
	#Step8:
	lineage_Name_LIST = ['Kingdom', 'Phylum', 'Class', 'Order', 'Family', 'Genus', 'Species']
	lineage_target_DICT = {}
	lineage_target_DICT['Kingdom'] = []
	lineage_target_DICT['Phylum'] = []
	lineage_target_DICT['Class'] = []
	lineage_target_DICT['Order'] = []
	lineage_target_DICT['Family'] = []
	lineage_target_DICT['Genus'] = []
	lineage_target_DICT['Species'] = []

	unique_taxonomy_LIST = []
	unique_taxonomy_LIST = list(set(TAX_DICT['TAXONOMY']))
	#print unique_taxonomy_LIST
	#'k_bacteria;p_bacteroidetes;c_bacteroidia;o_bacteroidales;f_bacteroidaceae;*f_bacteroidaceae;*f_bacteroidaceae;', 'k_bacteria;p_tenericutes;c_mollicutes;o_rf39;*o_rf39;*o_rf39;*o_rf39;', 'k_bacteria;p_firmicutes;c_clostridia;o_clostridiales;f_ruminococcaceae;*f_ruminococcaceae;*f_ruminococcaceae;',
	#'k_bacteria;p_firmicutes;c_clostridia;o_clostridiales;f_peptococcaceae;*f_peptococcaceae;*f_peptococcaceae;',
	for each_taxonomy in unique_taxonomy_LIST:
		taxonomy_LIST = each_taxonomy.split(';')
		lineage_target_DICT['Kingdom'].append(taxonomy_LIST[0])
		lineage_target_DICT['Phylum'].append(taxonomy_LIST[1])
		lineage_target_DICT['Class'].append(taxonomy_LIST[2])
		lineage_target_DICT['Order'].append(taxonomy_LIST[3])
		lineage_target_DICT['Family'].append(taxonomy_LIST[4])
		lineage_target_DICT['Genus'].append(taxonomy_LIST[5])
		lineage_target_DICT['Species'].append(taxonomy_LIST[6])
	for each_lineage in lineage_Name_LIST:
		lineage_target_DICT[each_lineage] = list(set(lineage_target_DICT[each_lineage]))
	#print lineage_target_DICT
	#{'Kingdom': ['k_Unclassified', 'k_Archaea', 'k_Bacteria'], 'Family': ['*o_Exiguobacterales', 'f_Fimbriimonadaceae'], 'Species': ['*o_RF39', '*f_Peptococcaceae',
	TAX_map_DICT = {}
	#alias_counter = 1
	TAX_temp_DICT = {}
	for each_lineage in lineage_Name_LIST:
		TAX_map_DICT[each_lineage] = {}
		for each_taxonomy in lineage_target_DICT[each_lineage]:
			TAX_ID = 'TAX_ID_' + str(TAX_ID_counter).zfill(6)
			TAX_DICT['TAX_ID'].append(TAX_ID)
			TAX_DICT['TAX_LEVEL'].append(each_lineage)
			TAX_DICT['TAXONOMY'].append(each_taxonomy)
			TAX_temp_DICT[each_lineage + ';;' + each_taxonomy] = TAX_ID
			TAX_map_DICT[each_lineage][TAX_ID] = []
			TAX_ID_counter += 1
	TAX_MATRIX = {}
	TAX_MATRIX = dict(zip(TAX_DICT['TAX_ID'], TAX_DICT['TAXONOMY']))
	Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(TAX_DICT, TAX_LIST, the_INDEX='TAX_ID')
	Pandas_DataFrame.to_excel(Excel_writer, sheet_name='TAX_metadata', columns=pandas_DataFrame_LIST, header=True, index=True)
	print "TAX_metadata created!!!"
	#print TAX_MATRIX
	#'TAX_ID_000485': '*f_peptococcaceae', 'TAX_ID_000188': 'k_bacteria;p_firmicutes;c_clostridia;o_clostridiales;
	#f_lachnospiraceae;*f_lachnospiraceae;*f_lachnospiraceae;', 'TAX_ID_000487': '*f_f16', 'TAX_ID_000486': '*f_coriobacteriaceae',
	# ######################################################################
	# TAXONOMY_INFER
	# ######################################################################

	for each_OTU in OTU_MATRIX_LIST:
		taxonomy_LIST = OTU_MATRIX[each_OTU].split(';')
		########################################
		TAX_map_DICT['Kingdom'][TAX_temp_DICT['Kingdom;;' + taxonomy_LIST[0]]].append(each_OTU)
		#############################################################################################################
		TAX_map_DICT['Phylum'][TAX_temp_DICT['Phylum;;' + taxonomy_LIST[1]]].append(each_OTU)
		#############################################################################################################
		TAX_map_DICT['Class'][TAX_temp_DICT['Class;;' + taxonomy_LIST[2]]].append(each_OTU)
		#############################################################################################################
		TAX_map_DICT['Order'][TAX_temp_DICT['Order;;' + taxonomy_LIST[3]]].append(each_OTU)
		#############################################################################################################
		TAX_map_DICT['Family'][TAX_temp_DICT['Family;;' + taxonomy_LIST[4]]].append(each_OTU)
		#############################################################################################################
		TAX_map_DICT['Genus'][TAX_temp_DICT['Genus;;' + taxonomy_LIST[5]]].append(each_OTU)
		#############################################################################################################
		TAX_map_DICT['Species'][TAX_temp_DICT['Species;;' + taxonomy_LIST[6]]].append(each_OTU)
		#############################################################################################################
	#print TAX_map_DICT['Kingdom']
	#{'TAX_ID_000420': ['OTU_ID_000343', 'OTU_ID_000342', 'OTU_ID_000101', 'OTU_ID_000100',
	for each_lineage in lineage_Name_LIST:
		#print each_lineage
		Lineage_DICT = {}
		Lineage_LIST = ['TAX_ID', 'TAXONOMY']
		Lineage_DICT['TAX_ID'] = []
		Lineage_DICT['TAXONOMY'] = []
		for each_TAX_ID in TAX_map_DICT[each_lineage]:
			Lineage_DICT['TAX_ID'].append(each_TAX_ID)
			Lineage_DICT['TAXONOMY'].append(TAX_MATRIX[each_TAX_ID])
		#Lineage_DICT['TAX_ID'].extend(TAX_map_DICT[each_lineage])
		for each_Sample in Biom_MATRIX_LIST:
			#print each_Sample
			Lineage_DICT[each_Sample] = []
			Lineage_LIST.append(each_Sample)
			for each_TAX_ID in TAX_map_DICT[each_lineage]:
				list_of_value = []
				for each_OTU_ID in TAX_map_DICT[each_lineage][each_TAX_ID]:
					#print each_OTU_ID
					list_of_value.append(Biom_MATRIX[each_Sample][each_OTU_ID])
				Lineage_DICT[each_Sample].append(sum(map(float, list_of_value)))
		Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(Lineage_DICT, Lineage_LIST, the_INDEX='TAX_ID')
		Pandas_DataFrame.to_excel(Excel_writer, sheet_name=each_lineage, columns=pandas_DataFrame_LIST, header=True, index=True)
		print each_lineage, " created!!!"
	# ######################################################################
	# OTU_LEVEL
	# ######################################################################
	OTU_level_DICT = {}
	OTU_level_LIST = ['TAX_ID', 'TAXONOMY']
	OTU_level_DICT['TAX_ID'] = []
	OTU_level_DICT['TAXONOMY'] = []
	for each_OTU in OTU_MATRIX_LIST:
		OTU_level_DICT['TAX_ID'].append(each_OTU.replace('OTU', 'TAX'))
		OTU_level_DICT['TAXONOMY'].append(OTU_MATRIX[each_OTU])
	for each_Sample in Sample_ID_LIST:
		OTU_level_DICT[each_Sample] = []
		OTU_level_LIST.append(each_Sample)
		OTU_level_DICT[each_Sample] = Biom_DICT[each_Sample]
	Pandas_DataFrame, pandas_DataFrame_LIST = dict_to_pandas_dataframe_converter(OTU_level_DICT, OTU_level_LIST, the_INDEX='TAX_ID')
	Pandas_DataFrame.to_excel(Excel_writer, sheet_name='OTU', columns=OTU_level_LIST, header=True, index=True)
	Excel_writer.save()
	print "OTU created!!!"
	#Emptying
	Biom_DICT = {}
	Lineage_DICT = {}
	TAX_map_DICT = {}
	TAX_MATRIX = {}
	TAX_temp_DICT = {}
	lineage_target_DICT = {}
	TAX_DICT = {}
	Design_DICT = {}
	OTU_DICT = {}
	OTU_MATRIX = {}
	Sample_DICT = {}
	Biom_MATRIX = {}
	Biom_DICT = {}
	OTU_level_DICT = {}
	return True


def dict_to_pandas_dataframe_converter(the_DICT, the_dict_header_LIST=None, the_INDEX=None):
	pd_DF = pandas.DataFrame.from_dict(the_DICT, orient='columns', dtype='float')
	if the_dict_header_LIST is not None:
		#sort the dataframe base on header list
		pd_DF = pd_DF[the_dict_header_LIST]
	if the_INDEX is not None:
		pd_DF = pd_DF.set_index(the_INDEX, drop=True, append=False, inplace=False, verify_integrity=False)
	if the_INDEX in the_dict_header_LIST:
		the_dict_header_LIST.remove(the_INDEX)
	return (pd_DF, the_dict_header_LIST)


def excel_to_vertical_dict_converter(Excel_file_PATH, SHEETname):
	DATAFRAME = pandas.read_excel(Excel_file_PATH, sheetname=SHEETname)
	any_vertical_LIST = []
	any_vertical_LIST = DATAFRAME.columns.values.tolist()
	any_vertical_DICT = {}
	for each_column in any_vertical_LIST:
		each_column_value_list = DATAFRAME[each_column].values.tolist()
		any_vertical_DICT[str(each_column)] = map(str, each_column_value_list)
	any_vertical_LIST = map(str, any_vertical_LIST)
	return (any_vertical_DICT, any_vertical_LIST)


def excel_to_pandas_dataframe_converter(Excel_File_PATH, SHEETname, the_INDEX=None):
	#
	return pandas.read_excel(Excel_File_PATH, sheetname=SHEETname, index_col=the_INDEX)


def text_to_pandas_dataframe_converter(Text_File_PATH, the_INDEX=None):
	# ##########
	return pandas.read_table(Text_File_PATH, low_memory=False, encoding='utf-8', skip_blank_lines=True, error_bad_lines=False, index_col=the_INDEX)


def excel_to_matrix_converter(Excel_file_PATH, SHEETname, index_col=None):
	if index_col is None:
		index_col = 0
	vertical_DICT, vertical_LIST = excel_to_vertical_dict_converter(Excel_file_PATH, SHEETname)
	Excel_MATRIX = {}
	Excel_LIST = vertical_LIST
	for each_SAMPLE in vertical_DICT[vertical_LIST[index_col]]:
		Excel_MATRIX[each_SAMPLE] = {}
		each_SAMPLE_index = vertical_DICT[vertical_LIST[index_col]].index(each_SAMPLE)
		for each_column in vertical_LIST:
			Excel_MATRIX[each_SAMPLE][each_column] = vertical_DICT[each_column][each_SAMPLE_index]
	return (Excel_MATRIX, Excel_LIST)


def excel_to_horizontal_dict_converter(Excel_file_PATH, SHEETname, index_col=None):
	DATAFRAME = pandas.read_excel(Excel_file_PATH, sheetname=SHEETname)
	if index_col is None:
		index_col = 0
	any_horizontal_DICT = {}
	any_horizontal_LIST = []
	for index, row in DATAFRAME.iterrows():
		row = map(str, row.tolist())
		any_horizontal_DICT[row[index_col]] = row
		any_horizontal_LIST.append(row[index_col])
	return (any_horizontal_DICT, any_horizontal_LIST)


def excel_to_shared_file_converter(Excel_File_PATH, SHEETname, Shared_File_PATH):
	#
	Excel_DataFrame = excel_to_pandas_dataframe_converter(Excel_File_PATH, SHEETname, 'TAX_ID')
	Shared_LIST = ['label', 'Groups', 'numOtus']
	TAX_ID_LIST = Excel_DataFrame.index.values.tolist()
	Shared_LIST.extend(TAX_ID_LIST)
	TAXONOMY_LIST = Excel_DataFrame['TAXONOMY'].values.tolist()
	TAX_ID_DICT = dict(zip(TAX_ID_LIST, TAXONOMY_LIST))
	Excel_DataFrame = Excel_DataFrame.drop('TAXONOMY', axis=1, level=None, inplace=False, errors='raise')
	#TRANSPOSE
	Transposed_DataFrame = Excel_DataFrame.T
	Groups_LIST = Transposed_DataFrame.index.values.tolist()
	NumOtus_value = str(len(TAX_ID_LIST))
	NumOtus_Column = [NumOtus_value] * len(Groups_LIST)
	Label_Column = [SHEETname] * len(Groups_LIST)
	Transposed_DataFrame = Transposed_DataFrame.assign(label=Label_Column)
	Transposed_DataFrame = Transposed_DataFrame.assign(Groups=Groups_LIST)
	Transposed_DataFrame = Transposed_DataFrame.assign(numOtus=NumOtus_Column)
	Transposed_DataFrame = Transposed_DataFrame[Shared_LIST]
	Transposed_DataFrame.to_csv(Shared_File_PATH, sep='\t', index=False, header=True)
	return TAX_ID_DICT


def any_dict_to_text_converter(any_DICT, any_LIST, text_file_PATH):
	if any_LIST is None:
		any_LIST = any_DICT.keys()
	pandas_DataFrame = pandas.DataFrame.from_dict(any_DICT, orient='columns')
	pandas_DataFrame.to_csv(text_file_PATH, sep='\t', columns=any_LIST, index=False, header=True)
	return True


def any_file_to_vertical_dict_converter(file_PATH):
	any_DATAFRAME = get_pandas_DATAFRAME(file_PATH)
	any_vertical_LIST = []
	any_vertical_LIST = any_DATAFRAME.columns.values.tolist()
	any_vertical_DICT = {}
	for each_column in any_vertical_LIST:
		each_column_value_list = any_DATAFRAME[each_column].values.tolist()
		any_vertical_DICT[str(each_column)] = map(str, each_column_value_list)
	any_vertical_LIST = map(str, any_vertical_LIST)
	return (any_vertical_DICT, any_vertical_LIST)
# ################################### CALCULATION ############################ #


def alpha_diversity_calculation(Excel_File_PATH, SHEETname, mothur_exec_PATH, processors_COUNT, outputdir_PATH):
	# ###############################
	Shared_File_PATH = outputdir_PATH + 'Shared_File.txt'
	TAX_ID_DICT = {}
	TAX_ID_DICT = excel_to_shared_file_converter(Excel_File_PATH, SHEETname, Shared_File_PATH)
	TAX_ID_DICT = {}
	# ######################## ALPHA DIVERSITY INDEX CALCULATION
	flag, stderr = execute_functions(mothur_summary_single, processors_COUNT, outputdir_PATH, 'multi', 'mothur', mothur_exec_PATH, Shared_File_PATH)
	if flag is False:
		print "Execution of summary_single failed!!!"
	else:
		scanned_container = []
		extension_list = ['.groups.summary']
		flag = scandirs(outputdir_PATH, scanned_container, extension_list)
		print "Scanning.."
		if flag is False:
			print "Failed :("
			print "This extension is not available: ", extension_list
			sys.exit(2)
		else:
			print "NICE :) Found it"
			counter = 1
			for file in scanned_container:
				print "File#", str(counter), ":", file
				counter += 1
	flag = remove_extension_files(outputdir_PATH, '.rabund')
	alpha_diversity_index_file = add_extension(scanned_container[0], '.txt')
	return alpha_diversity_index_file
# ################################### SPECIFIC ############################### #


def parse_mothur_logfile(logfile, keyword, output_container, mode=None):
	if mode == 'file':
		f = open(logfile, 'rU')
		data = f.read().split('\n')
	else:
		data = logfile
	for line in data.split('\n'):
		if keyword in line:
			output_container.append(line)
	if len(output_container) < 1:
		return False
	return True


def remove_ambiguity_from_taxonomy(taxonomy_LIST):
	#STEP 1: First remove any empty element in list
	taxonomy_LIST = filter(None, taxonomy_LIST)

	taxonomy_identifier = ['k_', 'p_', 'c_', 'o_', 'f_', 'g_', 's_']
	Clean_taxonomy_LIST = []

	#Step2: CLean up
	for each_tax in taxonomy_LIST:
		each_tax = str(each_tax).lower()
		each_tax = slugify(each_tax)
		if each_tax in ['root', 'unknown', 'domain', 'unclassified', 'unassigned', None, '', 'k_', 'p_', 'c_', 'o_', 'f_', 'g_', 's_']:
			continue
		elif 'unclassified' in each_tax:
			continue
		else:
			Clean_taxonomy_LIST.append(each_tax)
	if len(Clean_taxonomy_LIST) == 0:
		Clean_taxonomy_LIST = ['unclassified','unclassified','unclassified','unclassified','unclassified','unclassified','unclassified']

	last_index = 0
	tax_string = ''
	Last_tax = ''

	for each_level in taxonomy_identifier:
		each_level_index = taxonomy_identifier.index(each_level)
		try:
			Clean_taxonomy_LIST[each_level_index]
			last_index = each_level_index
			last_level = each_level
			if Clean_taxonomy_LIST[last_index][0] == '*':
				if Clean_taxonomy_LIST[last_index][1:3] in taxonomy_identifier:
					tax_string += Clean_taxonomy_LIST[last_index] + ';'
					Last_tax = Clean_taxonomy_LIST[last_index] + ';'
				else:
					tax_string += '*' + each_level + Clean_taxonomy_LIST[last_index][3:] + ';'
					Last_tax = '*' + each_level + Clean_taxonomy_LIST[last_index][3:] + ';'
			else:
				if Clean_taxonomy_LIST[last_index][0:2] in taxonomy_identifier:
					tax_string += Clean_taxonomy_LIST[last_index] + ';'
					Last_tax = Clean_taxonomy_LIST[last_index] + ';'
				else:
					tax_string += each_level + Clean_taxonomy_LIST[last_index] + ';'
					Last_tax = each_level + Clean_taxonomy_LIST[last_index] + ';'

		except IndexError:
			tax_string += '*' + Last_tax
	return tax_string


def any_dict_to_excel_file_append(Excel_File_PATH, sheetname, any_Dict, any_LIST=None):
	if any_LIST is None:
		any_LIST = any_Dict.keys()
	Excel_writer = pandas.ExcelWriter(Excel_File_PATH, engine='openpyxl')
	OpenBook = openpyxl.load_workbook(Excel_File_PATH)
	Excel_writer.book = OpenBook
	sheetnames_LIST = OpenBook.get_sheet_names()
	if sheetname in sheetnames_LIST:
		print "sheetname already exist"
		sys.exit(2)
	any_DICT_DataFrame = pandas.DataFrame.from_dict(any_Dict)
	any_DICT_DataFrame.to_excel(Excel_writer, sheet_name=sheetname, columns=any_LIST, index=False)
	Excel_writer.save()
	return True
# ################################### EXCEL VISUALIZER ####################### #


def alpha_diversity_in_excel(Excel_WorkBook, Alpha_DICT, Alpha_LIST, SHEETname=None):
	WorkSheet = Excel_WorkBook.create_sheet(SHEETname)
	Alpha_DF = pandas.DataFrame.from_dict(Alpha_DICT, orient='columns', dtype='float')
	Alpha_DF = Alpha_DF[Alpha_LIST]
	for each_Row in OPEN_DF.dataframe_to_rows(Alpha_DF, index=False, header=True):
		WorkSheet.append(tuple(each_Row))
	#Total_Number
	BarChart_OBJECT = OPEN_CH.BarChart()
	BarChart_OBJECT.type = "col"
	#BarChart_OBJECT.grouping = "stacked"
	BarChart_OBJECT.title = "Total read number bar chart"
	BarChart_OBJECT.style = 1
	BarChart_OBJECT.shape = 4
	#BarChart_OBJECT.overlap = 100
	BarChart_OBJECT.y_axis.title = 'Read Count'
	BarChart_OBJECT.x_axis.title = 'Sample name'
	#BarChart_OBJECT.width = 50
	#BarChart_OBJECT.height = 25
	Min_Col = 3
	Max_Col = 3
	Min_Row = 1
	Max_Row = len(Alpha_DICT['Sample_Name']) + 1

	Y_values = OPEN_CH.Reference(WorkSheet, min_col=Min_Col, min_row=Min_Row, max_col=Max_Col, max_row=Max_Row)
	X_values = OPEN_CH.Reference(WorkSheet, min_col=1, min_row=2, max_col=1, max_row=Max_Row)
	BarChart_OBJECT.add_data(Y_values, titles_from_data=True)
	BarChart_OBJECT.set_categories(X_values)

	WorkSheet.add_chart(BarChart_OBJECT, WorkSheet.cell(row=10, column=1).coordinate)

	#Classified_Unclassified
	BarChart_OBJECT = OPEN_CH.BarChart()
	BarChart_OBJECT.type = "col"
	BarChart_OBJECT.grouping = "stacked"
	BarChart_OBJECT.title = "Total read number bar chart"
	BarChart_OBJECT.style = 2
	BarChart_OBJECT.overlap = 100
	BarChart_OBJECT.y_axis.title = 'Read Count'
	BarChart_OBJECT.x_axis.title = 'Sample name'
	#BarChart_OBJECT.width = 50
	#BarChart_OBJECT.height = 25
	Min_Col = 4
	Max_Col = 5
	Min_Row = 1
	Max_Row = len(Alpha_DICT['Sample_Name']) + 1

	Y_values = OPEN_CH.Reference(WorkSheet, min_col=Min_Col, min_row=Min_Row, max_col=Max_Col, max_row=Max_Row)
	X_values = OPEN_CH.Reference(WorkSheet, min_col=1, min_row=2, max_col=1, max_row=Max_Row)
	BarChart_OBJECT.add_data(Y_values, titles_from_data=True)
	BarChart_OBJECT.set_categories(X_values)
	WorkSheet.add_chart(BarChart_OBJECT, WorkSheet.cell(row=10, column=Max_Row + 10).coordinate)

	#Diversity index
	BarChart_OBJECT = OPEN_CH.BarChart()
	BarChart_OBJECT.type = "col"
	BarChart_OBJECT.grouping = "standard"
	BarChart_OBJECT.title = "Diversity index"
	BarChart_OBJECT.style = 3
	#BarChart_OBJECT.overlap = 100
	BarChart_OBJECT.y_axis.title = 'Diversity value'
	BarChart_OBJECT.x_axis.title = 'Sample name'
	X_values = OPEN_CH.Reference(WorkSheet, min_col=1, min_row=2, max_col=1, max_row=Max_Row)
	#BarChart_OBJECT.width = 50
	#BarChart_OBJECT.height = 25
	MIN_COL = Alpha_LIST.index('Simpson_diversity_index') + 1
	MAX_COL = Alpha_LIST.index('Simpson_diversity_index') + 1
	MIN_ROW = 1
	MAX_ROW = len(Alpha_DICT['Sample_Name']) + 1
	Simpson_Value = OPEN_CH.Reference(WorkSheet, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW, max_row=MAX_ROW)
	Simpson_ErrorBars = error_bar_in_excel(Alpha_DICT['Simpson_High_Confidence_Interval'], Alpha_DICT['Simpson_Low_Confidence_Interval'])

	Simpson_Series = OPEN_SERIES.SeriesFactory(Simpson_Value, title_from_data=True)
	Simpson_Series.errBars = Simpson_ErrorBars
	BarChart_OBJECT.series.append(Simpson_Series)

	MIN_COL = Alpha_LIST.index('Shannon_diversity_index') + 1
	MAX_COL = Alpha_LIST.index('Shannon_diversity_index') + 1
	MIN_ROW = 1
	MAX_ROW = len(Alpha_DICT['Sample_Name']) + 1
	Shannon_Value = OPEN_CH.Reference(WorkSheet, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW, max_row=MAX_ROW)
	Shannon_ErrorBars = error_bar_in_excel(Alpha_DICT['Shannon_High_Confidence_Interval'], Alpha_DICT['Shannon_Low_Confidence_Interval'])

	Shannon_Series = OPEN_SERIES.SeriesFactory(Shannon_Value, title_from_data=True)
	Shannon_Series.errBars = Shannon_ErrorBars
	BarChart_OBJECT.series.append(Shannon_Series)

	BarChart_OBJECT.set_categories(X_values)
	WorkSheet.add_chart(BarChart_OBJECT, WorkSheet.cell(row=10, column=Max_Row + 20).coordinate)
	"""
	#MAX OTU ABUNDANCE
	BarChart_OBJECT = OPEN_CH.BarChart()
	BarChart_OBJECT.type = "col"
	BarChart_OBJECT.grouping = "standard"
	BarChart_OBJECT.title = "MAX OTU ABUNDANCE"
	BarChart_OBJECT.style = 3
	#BarChart_OBJECT.overlap = 100
	BarChart_OBJECT.y_axis.title = 'Otu Abundance'
	BarChart_OBJECT.x_axis.title = 'Sample name'
	for each_Sample in Alpha_DICT['Sample_Name']:
		each_Sample_index = Alpha_DICT['Sample_Name'].index(each_Sample) + 1
		X_values = OPEN_CH.Reference(WorkSheet, min_col=1, max_col=1, min_row=each_Sample_index, max_row=each_Sample_index)
	#BarChart_OBJECT.width = 50
	#BarChart_OBJECT.height = 25
		MIN_COL = Alpha_LIST.index('Maximum_Abundance_value') + 1
		MAX_COL = Alpha_LIST.index('Maximum_Abundance_value') + 1
		MIN_ROW = each_Sample_index
		MAX_ROW = each_Sample_index
		Label_COL = Alpha_LIST.index('Maximum_Abundance_OTU') + 1
		MAX_OTU_Value = OPEN_CH.Reference(WorkSheet, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW, max_row=MAX_ROW)
		MAX_OTU_Label = OPEN_CH.Reference(WorkSheet, min_col=Label_COL, max_col=Label_COL, min_row=MIN_ROW, max_row=MAX_ROW)
		MAX_OTU_Series = OPEN_SERIES.SeriesFactory(MAX_OTU_Value, title=MAX_OTU_Label, title_from_data=False)
		#BarChart_OBJECT.series.label = MAX_OTU_Label
		BarChart_OBJECT.series.append(MAX_OTU_Series)
	#BarChart_OBJECT.set_categories(X_values)
	WorkSheet.add_chart(BarChart_OBJECT, WorkSheet.cell(row=20, column=1).coordinate)

	#MIN OTU ABUNDANCE
	BarChart_OBJECT = OPEN_CH.BarChart()
	BarChart_OBJECT.type = "col"
	BarChart_OBJECT.grouping = "standard"
	BarChart_OBJECT.title = "MIN OTU ABUNDANCE"
	BarChart_OBJECT.style = 4
	#BarChart_OBJECT.overlap = 100
	BarChart_OBJECT.y_axis.title = 'Otu Abundance'
	BarChart_OBJECT.x_axis.title = 'Sample name'
	X_values = OPEN_CH.Reference(WorkSheet, min_col=1, min_row=2, max_col=1, max_row=Max_Row)

	MIN_COL = Alpha_LIST.index('Minimum_Abundance_value') + 1
	MAX_COL = Alpha_LIST.index('Minimum_Abundance_value') + 1
	MIN_ROW = 1
	MAX_ROW = len(Alpha_DICT['Sample_Name']) + 1
	MIN_OTU_Value = OPEN_CH.Reference(WorkSheet, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW, max_row=MAX_ROW)
	#Shannon_ErrorBars = error_bar_in_excel(Alpha_DICT['Shannon_High_Confidence_Interval'], Alpha_DICT['Shannon_Low_Confidence_Interval'])

	MIN_OTU_Series = OPEN_SERIES.SeriesFactory(MIN_OTU_Value, title_from_data=True)
	#Shannon_Series.errBars = Shannon_ErrorBars
	BarChart_OBJECT.series.append(MIN_OTU_Series)

	BarChart_OBJECT.set_categories(X_values)
	WorkSheet.add_chart(BarChart_OBJECT, WorkSheet.cell(row=20, column=Max_Row + 10).coordinate)
	"""
	return True


def rarefaction_in_excel(Excel_WorkBook, Data_DICT, Data_LIST, SHEETname=None):
	#Draw line chart using openpyxl
	WorkSheet = Excel_WorkBook.create_sheet('Rarefaction')
	DataFrame = pandas.DataFrame.from_dict(Data_DICT, orient='columns', dtype='float')
	DataFrame = DataFrame[Data_LIST]

	for each_Row in OPEN_DF.dataframe_to_rows(DataFrame, index=False, header=True):
		WorkSheet.append(each_Row)

	LineChart_OBJECT = OPEN_CH.LineChart()
	LineChart_OBJECT.title = "Rarefaction Curve Plot"
	LineChart_OBJECT.style = 7
	LineChart_OBJECT.y_axis.title = 'Detected Number of new species'
	LineChart_OBJECT.x_axis.title = 'Number of sequences sampled'
	LineChart_OBJECT.width = 50
	LineChart_OBJECT.height = 25
	Min_Col = 2
	Max_Col = len(Data_LIST)
	Min_Row = 1
	Max_Row = len(Data_DICT['x_axis']) + 1

	Y_values = OPEN_CH.Reference(WorkSheet, min_col=Min_Col, min_row=Min_Row, max_col=Max_Col, max_row=Max_Row)
	X_values = OPEN_CH.Reference(WorkSheet, min_col=1, min_row=2, max_col=1, max_row=Max_Row)
	LineChart_OBJECT.add_data(Y_values, titles_from_data=True)
	LineChart_OBJECT.set_categories(X_values)
	WorkSheet.add_chart(LineChart_OBJECT, "G10")
	return True


def natural_abundance_in_excel(Excel_WorkBook, Data_DICT, Data_LIST, SHEETname):
	WorkSheet = Excel_WorkBook.create_sheet(SHEETname)
	DataFrame = pandas.DataFrame.from_dict(Data_DICT, orient='columns', dtype='float')
	DataFrame = DataFrame[Data_LIST]
	for each_Row in OPEN_DF.dataframe_to_rows(DataFrame, index=False, header=True):
		WorkSheet.append(tuple(each_Row))

	BarChart_OBJECT = OPEN_CH.BarChart()
	BarChart_OBJECT.type = "col"
	BarChart_OBJECT.title = "Natural abundance BarPlot"
	BarChart_OBJECT.style = 7
	BarChart_OBJECT.y_axis.title = 'Entities abundance'
	BarChart_OBJECT.x_axis.title = 'Sample name'
	BarChart_OBJECT.width = 50
	BarChart_OBJECT.height = 25
	Min_Col = 2
	Max_Col = len(Data_LIST)
	Min_Row = 1
	Max_Row = len(Data_DICT['TAX_ID']) + 1

	Y_values = OPEN_CH.Reference(WorkSheet, min_col=Min_Col, min_row=Min_Row, max_col=Max_Col, max_row=Max_Row)
	X_values = OPEN_CH.Reference(WorkSheet, min_col=1, min_row=2, max_col=1, max_row=Max_Row)
	BarChart_OBJECT.add_data(Y_values, titles_from_data=True)
	BarChart_OBJECT.set_categories(X_values)
	WorkSheet.add_chart(BarChart_OBJECT, "G10")
	return True


def error_bar_in_excel(plus_LIST, minus_LIST):
	errDir = 'y'
	errValType = 'cust'
	#Convert to list of NumVal
	numvals_plus = [OPEN_DS.NumVal(i, None, v=x) for i, x in enumerate(plus_LIST)]
	numvals_minus = [OPEN_DS.NumVal(i, None, v=x) for i, x in enumerate(minus_LIST)]

	# Convert to NumData
	nd_plus = OPEN_DS.NumData(pt=numvals_plus)
	nd_minus = OPEN_DS.NumData(pt=numvals_minus)

	# Convert to NumDataSource
	nds_plus = OPEN_DS.NumDataSource(numLit=nd_plus)
	nds_minus = OPEN_DS.NumDataSource(numLit=nd_minus)

	ErrorBars_OBJECT = OPEN_CH.error_bar.ErrorBars(plus=nds_plus, minus=nds_minus, errDir=errDir, errValType=errValType)
	return ErrorBars_OBJECT
# ################################### PLOTLY VISUALIZER ###################### #


def html_visualizer(final_string, html_string=None, javascript_string=None):

	if html_string is None and javascript_string is None:
		if final_string == '':

			print "Initial html script appending"
			final_string += """
				<!DOCTYPE html>
							<html lang='en' xml:lang='en' xmlns='http://www.w3.org/1999/xhtml'>
								<head>
									<meta charset="utf-8">
									<meta name="viewport" content="width=device-width, initial-scale=1">
									<!--JQUERY LIBRARAY MAIN LIBRARY-->
									<script type="text/javascript" src="https://code.jquery.com/jquery-1.12.3.js"></script>
									<!--PLOTLY LIBRARAY MAIN LIBRARY-->
									<script type="text/javascript" async src="https://cdn.mathjax.org/mathjax/latest/MathJax.js?config=TeX-AMS-MML_SVG"></script>
									<script type="text/javascript" src="https://cdn.plot.ly/plotly-latest.min.js"></script>
									<!--BOOTSTRAP LIBRARAY MAIN LIBRARY-->
									<script type="text/javascript" src="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.7/js/bootstrap.min.js"></script>
									<link rel="stylesheet" type="text/css" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.7/css/bootstrap.min.css"/>
									<!--BOOTSTRAP DATATABLES LIBRARAY MAIN LIBRARY-->
									<link rel="stylesheet" type="text/css" href="https://cdn.datatables.net/1.10.12/css/dataTables.bootstrap.min.css"/>
									<script type="text/javascript" src="https://cdn.datatables.net/1.10.12/js/jquery.dataTables.min.js"></script>
									<script type="text/javascript" src="https://cdn.datatables.net/1.10.12/js/dataTables.bootstrap.min.js"></script>
									<!--FONTAWESOME LIBRARAY MAIN LIBRARY-->
									<link rel="stylesheet" type="text/css" href="ttps://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css"/>
									<!--############################-->

									<!--OPEN SCRIPTS-->
										<script type="text/javascript">
											$(document).ready(function() {
												$('#statistics_table').DataTable( {
													"order": [[ 0, "asc" ]],
													"paging": false,
													"info": false,
													"ordering": true,
													"searching": false,
													//"scrollY": "600px",
													"lengthMenu": [[25, 50, -1], [25, 50, "All"]]
												});
												$('#biomarker_table').DataTable( {
													"order": [[ 1, "asc" ]],
													"paging": false,
													"info": false,
													"ordering": true,
													"searching": false,
													"scrollY": "600px",
													"lengthMenu": [[25, 50, -1], [25, 50, "All"]]
												});

											});
									//###################################################
									//#######RELOAD ON RESIZE
									window.addEventListener('resize', function () {
									"use strict";
									window.location.reload();
									});

									//###################################################
											function linechart_filter(target_div){
											var myPlot = document.getElementById(target_div);
											var new_data = [];
											myPlot.on('plotly_click', function(targetdata){
											Plotly.deleteTraces(myPlot, targetdata.points[0].curveNumber)
												});

											};
									//###################################################
											function barchart_filter(target_div){
											var myPlot = document.getElementById(target_div);
											var new_data = [];
											myPlot.on('plotly_click', function(targetdata){

											var index_value = '';
											index_value = targetdata.points[0].x;
											var index_place = 0;
											index_place  = targetdata.points[0].fullData.x.indexOf(index_value);
											//console.log(index_value);
											//console.log(index_place);
											//console.log(targetdata);
											main_data_size = myPlot.data.length;
											new_data = myPlot.data
											for (var main_data_index = 0; main_data_index < main_data_size; main_data_index++){
												new_data[main_data_index].x.splice(index_place, 1);
												new_data[main_data_index].y.splice(index_place, 1);

											}
											//console.log(data_sample_abundance_Cancer[0].x)
											//Plotly.newPlot('myDiv', data, layout);
											//Plotly.deleteTraces(myPlot, data.points[0].x)
											//Plotly.newPlot('myDiv', data, layout);
											Plotly.redraw(target_div, new_data)
											//Plotly.relayout('myDiv',data)
										});

										};
										//##############################################
										function add_annotate(target_div){
											var myPlot = document.getElementById(target_div);
											var new_data = [];
											myPlot.on('plotly_click', function(target_data){
												var index_value = '';
												index_value = target_data.points[0].x;
												var index_place = 0;
												index_place  = target_data.points[0].fullData.x.indexOf(index_value);
											//console.log(index_value);
											//console.log(index_place);
											//console.log(targetdata);
												main_data_size = myPlot.data.length;
												new_data = myPlot.data;
												var annotation_string = 'Max:';
												var abundance_value_dict = {};
												for (var main_data_index = 0; main_data_index < main_data_size; main_data_index++){
													abundance_value_dict[myPlot.data[main_data_index].name] = myPlot.data[main_data_index].y[index_place];
												}
												sorted_abundance_value_array = sort_dictionary_by_value_descendingly(abundance_value_dict);
												//console.log(sorted_abundance_value_array);
												yPosition = 0;
												for (var i = 0; i < 1; i++){

													annotation_string += sorted_abundance_value_array[i] + '(' + abundance_value_dict[sorted_abundance_value_array[i]] + ')';

												}
												console.log(target_data.points[0].fullData.y);
												yPosition = target_data.points[0].fullData.y.reduce((a, b) => a + b, 0);
												console.log(yPosition);
												annotation = {
													text: annotation_string,
													x: index_value,
													y: yPosition,
													showarrow: true,
													font: {
														family: 'Avenir',
														size: 10,
														//color: '#ffffff'
													},
													align: 'center',
													arrowhead: 2,
													arrowsize: 1,
													arrowwidth: 2,
													arrowcolor: '#636363',
													ax: 20,
													ay: -30,
													bordercolor: '#c7c7c7',
													borderwidth: 2,
													borderpad: 4,
													//bgcolor: '#ff7f0e',
													opacity: 0.8
												}
												annotations = myPlot.layout.annotations || [];
												annotations.push(annotation);
												Plotly.relayout(target_div,{annotations: annotations});

											});
										};
										//##############################################
										function add_comment(target_div, target_comment){
											var myPlot = document.getElementById(target_div);
											var myComment = document.getElementById(target_comment).value;
											console.log(myComment);
											var new_data = [];
											myPlot.on('plotly_click', function(target_data){
												var index_value = '';
												index_value = target_data.points[0].x;
												var index_place = 0;
												index_place  = target_data.points[0].fullData.x.indexOf(index_value);
												annotation = {
													text: myComment,
													x: index_value,
													//y: yPosition,
													showarrow: true,
													font: {
														family: 'Avenir',
														size: 10,
														//color: '#ffffff'
													},
													align: 'center',
													arrowhead: 2,
													arrowsize: 1,
													arrowwidth: 2,
													arrowcolor: '#636363',
													ax: 20,
													ay: -30,
													bordercolor: '#c7c7c7',
													borderwidth: 2,
													borderpad: 4,
													//bgcolor: '#ff7f0e',
													opacity: 0.8
												}
												annotations = myPlot.layout.annotations || [];
												annotations.push(annotation);
												Plotly.relayout(target_div,{annotations: annotations});

											});
										};
										//##############################################

										function sort_dictionary_by_value_descendingly(myDict){
											var keys = [];
											for(var key in myDict){
												keys.push(key);
											}
											return keys.sort(function(a,b){return myDict[b] - myDict[a]});

										};
										</script>
									<!--############################-->
									<!--STYLES-->
									<style type="text/css">
										body {
										position: relative;
										}
										.table-hover tbody tr:hover td, .table-hover tbody tr:hover th {
														background-color: yellow;
													}
										a:link {color: #669;}		/* unvisited link */
										a:visited {color: #669;}	/* visited link */
										a:hover {color: #669;}		/* mouse over link */
										a:active {color: #669;}		/* selected link */
										#design {padding-top:50px;}
										#alpha_diversity {padding-top:50px;}
										#rarefaction {padding-top:50px;}
										#sample_abundance {padding-top:50px;}
										#bacterial_abundance {padding-top:50px;}
										#pcoa {padding-top:50px;}
										#biomarker_discovery {padding-top:50px;}
									</style>
									<!--############################-->
								</head>
								<body>
								<div class="container-fluid">
								<h2>BiomeSlicer results</h2>
								<div class="row">
									<dl class="dl-horizontal">
									<dt>Request Name</dt>
									<dd>20160216-H1</dd>
									<dt>Project Title</dt>
									<dd>biom composition</dd>
									<dt>Project Description</dt>
									<dd>
									<strong> Brief background: </strong><small>Breif background about projects</small><br>
									<strong>Goal/purpose or rationale: </strong><small>Describe why are doing this</small></dd-->
									<dt>Experiment</dt>
									<dd>Experiments value</dd>
									<dt>PI</dt>
									<dd>...</dd>
									</dl>
								</div>
				"""
		else:
			final_string += '</body>\n</html>\n'
	elif html_string is not None and javascript_string is None:
		final_string += html_string
	elif html_string is not None and javascript_string is not None:
		final_string += html_string
		final_string += javascript_string
	return final_string


def plotly_html_maker(main_div_name, plot_type, figure_object_dict, mode_bar=None):
	temp_plotly_string = ''
	for each_figure in figure_object_dict:
		print each_figure
		plotly_div = plotly.offline.plot(figure_object_dict[each_figure], include_plotlyjs=False, output_type='div', show_link=False)
		temp_string = plotly_div.split('Plotly.newPlot(', 1)[1]
		temp2_string = temp_string.split(',', 1)[1]
		plotly_javascript_string = temp2_string.split('})', 1)[0]
		if mode_bar is None:
			plotly_javascript_string += """, displaylogo: false, displayModeBar: true, modeBarButtonsToRemove: ['zoom2d', 'select2d', 'zoomIn2d', 'zoomOut2d', 'autoScale2d', 'sendDataToCloud', 'orbitRotation', 'tableRotation', 'pan3d', 'zoom3d','resetCameraDefault3d','resetCameraLastSave3d','hoverClosest3d'],}"""
		elif mode_bar is False:
			plotly_javascript_string += """, displaylogo: false, displayModeBar: false, modeBarButtonsToRemove: ['zoom2d', 'select2d', 'zoomIn2d', 'zoomOut2d', 'autoScale2d', 'sendDataToCloud', 'orbitRotation', 'tableRotation', 'pan3d', 'zoom3d','resetCameraDefault3d','resetCameraLastSave3d','hoverClosest3d'],}"""
		temp_plotly_string += """Plotly.newPlot("PLOTLY_""" + main_div_name + '_' + plot_type + '_' + each_figure + """_DIV", """ + plotly_javascript_string + """);"""

	panel_count = len(figure_object_dict.keys())
	if panel_count == 1:
		bootstrap_column_string = 'col-xs-12 col-sm-12 col-md-12 col-lg-12'
	if panel_count == 2:
		bootstrap_column_string = 'col-xs-6 col-sm-6 col-md-6 col-lg-6'
	elif panel_count == 3:
		bootstrap_column_string = 'col-xs-4 col-sm-4 col-md-4 col-lg-4'
	elif panel_count == 4:
		bootstrap_column_string = 'col-xs-3 col-sm-3 col-md-3 col-lg-3'
	elif panel_count > 4:
		bootstrap_column_string = 'col-xs-2 col-sm-2 col-md-2 col-lg-2'

	plotly_div_string = """
		<div class="row">
	"""

	for each_figure in figure_object_dict:
		plotly_div_string += """
					<div id="PLOTLY_""" + main_div_name + '_' + plot_type + '_' + each_figure + """_DIV" class=" """ + bootstrap_column_string + """ "></div>

		"""

	plotly_div_string += """
		</div>
	"""

	plotly_html_string = """
	<h4>""" + main_div_name + ' ' + plot_type + """</h4>
	<div class="row">
		<dl class="dl-horizontal">
		<dt>Brief Description</dt>
		<dd>
		<strong> Brief background: </strong><small>Some description over here is useful</small><br>
		<strong>Goal/purpose or rationale: </strong><small>Alittle detail of the purpose over here</small></dd-->
		</dl>
	</div>
	<div id="PLOTLY_""" + main_div_name + '_' + plot_type + """_MAIN_DIV" class="container-fluid">
		<div class="row">

			<div class="col-xs-12 col-sm-12 col-md-12 col-lg-12">
				<div class="panel panel-default" >
					<div id="PLOTLY_""" + main_div_name + '_' + plot_type + """_HEADER_DIV" class="panel-heading" style="border:0px; background-color: #F4F9FD;">
						<h3 class="panel-title">""" + main_div_name + """</h3>
					</div>
					<div id="PLOTLY_""" + main_div_name + '_' + plot_type + """_BODY_DIV" class="panel-body" align="center">

							""" + plotly_div_string + """
					</div>
					<div id="PLOTLY_""" + main_div_name + '_' + plot_type + """_FOOTER_DIV" class="panel-footer">
						<form class="form-inline" role="form">
							<label class="form-check-inline" style="font-family:'Avenir';margin-left: 5px;">
								<input type="radio" name="options" id="ALPHA_DIVERSITY_INDEX_DIV_normal_mode" autocomplete="off" checked > Normal mode
							</label>
							<!--label class="form-check-inline" style="font-family:'Avenir';margin-left: 5px;">
								<input type="radio" name="options" id="ALPHA_DIVERSITY_INDEX_DIV_delete_sample_mode" autocomplete="off" onclick="barchart_filter('ALPHA_DIVERSITY_INDEX_DIV');"> Click to Delete mode
							</label-->
						</form>
					</div>
				</div>
			</div>
		</div>
	</div>
	"""
	plotly_script = """
				<script type="text/javascript">
					//####################################################################
					//####################################################################
					//####################################################################
					//#####   PLOTLY_""" + main_div_name + '_' + plot_type + """  ########
					//####################################################################
					//####################################################################
					//####################################################################
					//####################################################################
	"""
	plotly_script += temp_plotly_string

	plotly_script += """
					//####################################################################
					//####################################################################
					//####################################################################
				</script>
	"""
	return (plotly_html_string, plotly_script)


def rarefaction_curve_plot(Excel_File_PATH, mothur_exec_path, processors, outputdir_PATH):
	#Generate Shared_file
	Shared_File_PATH = excel_to_shared_file_converter(Excel_File_PATH, outputdir_PATH)
	##################################################################################################################
	Shared_vertical_DICT, Shared_vertical_LIST = any_file_to_vertical_dict_converter(Shared_File_PATH)
	#1 heuristic method to find the frequecy value value
	OTU_total_count_LIST = []
	for each_OTU in Shared_vertical_LIST[3:]:
		OTU_total_count_LIST.append(sum(map(int, Shared_vertical_DICT[each_OTU])))
	min_threshold = 50
	#print OTU_total_count_LIST
	rare_value = numpy_percentile(OTU_total_count_LIST, min_threshold)
	#sys.exit(2)
	#2 remove rare
	frequency_value = str(int(rare_value))
	#Run RAREFACTION
	# ######################## Rarefaction curve CALCULATION
	#frequency_value = '0.25'
	flag, stderr = execute_functions(mothur_rarefaction_single, processors, outputdir_PATH, 'multi', 'mothur', mothur_exec_path, Shared_File_PATH, frequency_value)
	if flag is False:
		print "Execution of mothur_rarefaction_single failed!!!"
		sys.exit(2)
	else:
		scanned_container = []
		extension_list = ['.groups.r_chao']
		flag = scandirs(outputdir_PATH, scanned_container, extension_list)
		print "Scanning.."
		if flag is False:
			print "Failed :("
			print "This extension is not available: ", extension_list
			sys.exit(2)
		else:
			print "NICE :) Found it"
			counter = 1
			for file in scanned_container:
				print "File#", str(counter), ":", file
				counter += 1
	flag = remove_extension_files(outputdir_PATH, '.rabund')
	rarefaction_file = add_extension(scanned_container[0], '.txt')
	##################################################################################################################
	# MANAGING DESIGN AND METADATA
	SAMPLE_metadata_MATRIX, SAMPLE_metadata_LIST = excel_to_matrix_converter(Excel_File_PATH, 'Sample_metadata', 0)
	if DESIGN_EXIST is True:
		Design_MATRIX, Design_LIST = excel_to_matrix_converter(Excel_File_PATH, 'Design', 0)
	# ######################## PARSING
	rarefaction_file_vertical_DICT, rarefaction_file_vertical_LIST = any_file_to_vertical_dict_converter(rarefaction_file)
	number_of_sequences_sampled_list = rarefaction_file_vertical_DICT['numsampled']
	
	raw_sample_name_list = []
	for each_head in rarefaction_file_vertical_LIST:
		if 'lci' in each_head:
			continue
		elif 'hci' in each_head:
			continue
		elif 'numsampled' in each_head:
			continue
		else:
			raw_sample_name_list.append(each_head)

	rarefaction_sample_dict_values = {}
	for each_sample in raw_sample_name_list:
		clean_sample_name = each_sample.split('-')[1]
		rarefaction_sample_dict_values[clean_sample_name] = list_variance(rarefaction_file_vertical_DICT[each_sample], 0.01)  # removing NA element and variant factors
	# ######################## Create VISIBILITY DICT
	visibility_list_length = len(rarefaction_sample_dict_values.keys())
	visibility_flag_dict = {}
	visibility_flag_dict['ALL'] = [True] * visibility_list_length
	VISIBLE_flag_list = []
	VISIBLE_flag_list.append(
		dict(
			args=[
				dict(
					#title=title_dictionary[each_index],
					visible=visibility_flag_dict['ALL']
				)
			],
			label='ALL',
			method='restyle',

		),
	)
	# ######################## PLOTLY DATA CREATE
	RAREFACTION_curve_linechart_TRACE_list = []
	RAREFACTION_curve_linechart_LAYOUT_objects_dict = {}
	RAREFACTION_curve_linechart_FIGURE_objects_dict = {}
	for each_sample in rarefaction_sample_dict_values.keys():
		RAREFACTION_curve_linechart_TRACE_list.append(PLOTLY_GO.Scatter(
				x=number_of_sequences_sampled_list,
				y=rarefaction_sample_dict_values[each_sample],
				name=SAMPLE_metadata_MATRIX[each_sample]['Sample_ID'],

				hoverinfo='name',
				mode='lines',
			)
		)
	RAREFACTION_curve_linechart_LAYOUT_objects_dict = PLOTLY_GO.Layout(
		height=600,
		autosize=True,
		title='Rarefaction curve plot',
		showlegend=True,
		hovermode='closest',
		titlefont=dict(
			family='Avenir',
			size=12
		),
		xaxis=dict(
			showgrid=False,
			title='Number of sampling effort',
			titlefont=dict(
				family='Avenir',
				size=10,
			)

		),
		font=dict(family='Avenir', size=10),
		yaxis=dict(
			showgrid=True,
			title='Number of detected species',
			titlefont=dict(
				family='Avenir',
				size=10,
			)
		),
		margin=dict(
			l=100,
			r=100,
			b=100,
			t=100,
		),
		updatemenus=list([
			dict(
				direction="down",
				y=1,
				x=-0.05,
				font=dict(family='Avenir', size=12),
				buttons=VISIBLE_flag_list,
				type="buttons",
				active=0,
				visible=True
			)
		]
		)
	)
	RAREFACTION_curve_linechart_FIGURE_objects_dict['rarefaction'] = PLOTLY_GO.Figure(data=RAREFACTION_curve_linechart_TRACE_list, layout=RAREFACTION_curve_linechart_LAYOUT_objects_dict)
	# ###################################  PLOTLY PLOT CREATE
	plotly_html_string, plotly_script = plotly_html_maker('RAREFACTION_CURVE', 'LINE_CHART', RAREFACTION_curve_linechart_FIGURE_objects_dict, mode_bar=False)
	return (plotly_html_string, plotly_script)


def rarefaction_in_html(FINAL_STRING, Data_DICT, Data_LIST):
	LineChart_TRACE_OBJECTS_LIST = []
	LineChart_LAYOUT_OBJECTS = {}
	LineChart_FIGURE_OBJECTS = {}
	X_values = Data_DICT['x_axis']
	for each_Column in Data_LIST[1:]:
		LineChart_TRACE_OBJECTS_LIST.append(PLOTLY_GO.Scatter(
				x=X_values,
				y=Data_DICT[each_Column],
				name=each_Column,
				hoverinfo='name',
				mode='lines',
			)
		)
	LineChart_LAYOUT_OBJECTS = PLOTLY_GO.Layout(
		height=600,
		autosize=True,
		title='Rarefaction curve plot',
		showlegend=True,
		hovermode='closest',
		titlefont=dict(
			family='Avenir',
			size=12
		),
		xaxis=dict(
			showgrid=False,
			title='Number of sampling effort',
			titlefont=dict(
				family='Avenir',
				size=10,
			)

		),
		font=dict(family='Avenir', size=10),
		yaxis=dict(
			showgrid=True,
			title='Number of detected species',
			titlefont=dict(
				family='Avenir',
				size=10,
			)
		),
		margin=dict(
			l=100,
			r=100,
			b=100,
			t=100,
		)
	)
	LineChart_FIGURE_OBJECTS['Rarefaction'] = PLOTLY_GO.Figure(data=LineChart_TRACE_OBJECTS_LIST, layout=LineChart_LAYOUT_OBJECTS)
	plotly_html_STRING, plotly_javascript_STRING = plotly_html_maker('RAREFACTION_CURVE', 'LINE_CHART', LineChart_FIGURE_OBJECTS, mode_bar=False)
	FINAL_STRING = html_visualizer(FINAL_STRING, plotly_html_STRING, plotly_javascript_STRING)
	return FINAL_STRING


def alpha_diversity_in_html(FINAL_STRING, Alpha_DICT, Alpha_LIST):
	ALPHA_TRACE_OBJECTS_LIST = []
	ALPHA_LAYOUT_OBJECTS = {}
	ALPHA_FIGURE_OBJECTS = {}
	ALPHA_TRACE_OBJECTS_LIST.append(PLOTLY_GO.Bar(
			x=Alpha_DICT['Sample_Name'],
			y=Alpha_DICT['Total_Number'],
			hoverinfo='name',
		)
	)
	ALPHA_LAYOUT_OBJECTS = PLOTLY_GO.Layout(
		barmode='group',
		height=600,
		autosize=True,
		title='Relative abundance',
		showlegend=True,
		hovermode='closest',
		titlefont=dict(
			family='Avenir',
			size=16
			#color='#7f7f7f'
		),
		font=dict(family='Avenir', size=10),
		xaxis=dict(
			autorange=True,
		),
		yaxis=dict(
			title='Relative abundance',
			titlefont=dict(family='Avenir', size=12),
		),
	)
	ALPHA_FIGURE_OBJECTS['Alpha'] = PLOTLY_GO.Figure(data=ALPHA_TRACE_OBJECTS_LIST, layout=ALPHA_LAYOUT_OBJECTS)
	plotly_html_STRING, plotly_javascript_STRING = plotly_html_maker('alpha', 'BAR_CHART', ALPHA_FIGURE_OBJECTS, mode_bar=False)
	FINAL_STRING = html_visualizer(FINAL_STRING, plotly_html_STRING, plotly_javascript_STRING)
	return FINAL_STRING

# ################################### MOTHUR FUNCTIONS ####################### #


def test_mothur(processors, outputdir, stderr, stdout, run_pid, mothur_exec_PATH):
	# test mothur to see if it is working and grab the version of mothur by scanning its output log
	mothur_input_dictionary = {}
	command = 'get.current'
	mothur_input_dictionary['command'] = command
	mothur_input_dictionary['mothur_exec_path'] = mothur_exec_PATH
	mothur_input_dictionary['processors'] = processors
	mothur_input_dictionary['outputdir'] = outputdir
	mothur_input_dictionary['nohup_in'] = 'nohup'
	mothur_input_dictionary['nohup_out'] = '> ' + stdout + ' 2> ' + stderr + ' & echo $! > ' + run_pid
	#parameter_list = []
	make_file_object = mothur_process(mothur_input_dictionary)
	exec_dict = {}
	flag, exec_dict = make_file_object.execute_mothur_command()
	if flag is False:
		print "[FATAL-ERROR]: Commandline can not be executed!!!"
		print "ABORTING!!!"
		sys.exit(2)
	return True


def mothur_rarefaction_single(processors, outputdir, stderr, stdout, run_pid, mothur_exec_path, shared_file, freq_value=None):
	# grab path of paired fastq files and save it into container
	space = ' '
	mothur_input_dictionary = {}
	command = 'rarefaction.single'
	mothur_input_dictionary['command'] = command
	mothur_input_dictionary['mothur_exec_path'] = mothur_exec_path
	mothur_input_dictionary['processors'] = processors
	mothur_input_dictionary['outputdir'] = outputdir
	mothur_input_dictionary['nohup_in'] = 'nohup'
	mothur_input_dictionary['nohup_out'] = '> ' + stdout + ' 2> ' + stderr + ' & echo $! > ' + run_pid
	parameter_list = []
	parameter_list.append('shared=' + shared_file)
	parameter_list.append(',' + space + 'iters=100, calc=chao, groupmode=T')
	if freq_value is not None:
		parameter_list.append(',' + space + 'freq=' + str(freq_value))
	mothur_input_dictionary['parameters'] = parameter_list
	make_file_object = mothur_process(mothur_input_dictionary)
	exec_dict = {}
	flag, exec_dict = make_file_object.execute_mothur_command()
	if flag is False:
		print "[FATAL-ERROR]: Commandline can not be executed!!!"
		print "ABORTING!!!"
		sys.exit(2)
	return True


def mothur_summary_single(processors, outputdir, stderr, stdout, run_pid, mothur_exec_path, shared_file):
	# grab path of paired fastq files and save it into container
	space = ' '
	mothur_input_dictionary = {}
	command = 'summary.single'
	mothur_input_dictionary['command'] = command
	mothur_input_dictionary['mothur_exec_path'] = mothur_exec_path
	mothur_input_dictionary['processors'] = processors
	mothur_input_dictionary['outputdir'] = outputdir
	mothur_input_dictionary['nohup_in'] = 'nohup'
	mothur_input_dictionary['nohup_out'] = '> ' + stdout + ' 2> ' + stderr + ' & echo $! > ' + run_pid
	parameter_list = []
	parameter_list.append('shared=' + shared_file)
	parameter_list.append(',' + space + 'size=100000, iters=10000, calc=simpson-shannon-jack')
	mothur_input_dictionary['parameters'] = parameter_list
	make_file_object = mothur_process(mothur_input_dictionary)
	exec_dict = {}
	flag, exec_dict = make_file_object.execute_mothur_command()
	if flag is False:
		print "[FATAL-ERROR]: Commandline can not be executed!!!"
		print "ABORTING!!!"
		sys.exit(2)
	return True
# ################################### ANALYSIS ############################# #


def alpha_diversity_analysis(Excel_File_PATH, Excel_WorkBook, FINAL_STRING, mothur_exec_PATH, processors_COUNT, outputdir_PATH):
	# ###############################
	#First alpha diversity calculations
	Alpha_Diversity_File_PATH = alpha_diversity_calculation(Excel_File_PATH, 'OTU', mothur_exec_PATH, processors_COUNT, outputdir_PATH)
	Alpha_DF = text_to_pandas_dataframe_converter(Alpha_Diversity_File_PATH, the_INDEX='group')
	Alpha_Columns = Alpha_DF.columns.values.tolist()
	Alpha_Sample_Name_LIST = Alpha_DF.index.values.tolist()
	# ###############################
	# LOADING OTU DATAFRAME
	OTU_DataFrame = excel_to_pandas_dataframe_converter(Excel_File_PATH, 'OTU', the_INDEX='TAX_ID')
	TAX_ID_LIST = OTU_DataFrame.index.values.tolist()
	TAXONOMY_LIST = OTU_DataFrame['TAXONOMY'].values.tolist()
	TAX_ID_DICT = {}
	TAX_ID_DICT = dict(zip(TAX_ID_LIST, TAXONOMY_LIST))
	# Remove 'TAXONOMY' Column
	OTU_DataFrame = OTU_DataFrame.drop('TAXONOMY', axis=1, level=None, inplace=False, errors='raise')
	Sample_name_LIST = OTU_DataFrame.columns.values.tolist()
	# Look for unclassified entities
	Unclassified_Flag = False
	unclassified_index_LIST = []
	for each_OTU, each_TAX in itertools.izip(TAXONOMY_LIST, TAX_ID_LIST):
		if 'unclassified' in each_OTU:
			unclassified_index_LIST.append(each_TAX)
			Unclassified_Flag = True
	# ##############################
	# LOADING SAMPLE METADATA DATAFRAME
	Sample_metadata_DataFrame = excel_to_pandas_dataframe_converter(Excel_File_PATH, 'Sample_metadata', the_INDEX='Sample_ID')
	Sample_ID_LIST = Sample_metadata_DataFrame.index.values.tolist()
	Sample_metadata_Columns = Sample_metadata_DataFrame.columns.values.tolist()
	# #############################
	# Doing the actual process
	Alpha_Diversity_DICT = {}
	Alpha_Diversity_LIST = []
	Alpha_Diversity_LIST.extend(Sample_metadata_Columns)
	Alpha_Diversity_LIST.extend([
		'Total_Number',
		'Classified_Number',
		'Unclassified_Number',
		'Unclassified_Percentage',
		'Classified_Percentage',
		'Mean_Value',
		'Simpson_diversity_index',
		'Simpson_High_Confidence_Interval',
		'Simpson_Low_Confidence_Interval',
		'Shannon_diversity_index',
		'Shannon_High_Confidence_Interval',
		'Shannon_Low_Confidence_Interval',
		'Maximum_Abundance_value',
		'Minimum_Abundance_value',
		'Maximum_Abundance_OTU',
		'Minimum_Abundance_OTU'
	])
	
	###############################################################################
	table_head_string = ''
	table_head_string += '<thead>\n'
	table_head_string += '\t<tr>\n'
	for each_Alpha in Alpha_Diversity_LIST:
		Alpha_Diversity_DICT[each_Alpha] = []
		table_head_string += '\t\t<th class="text-center">' + each_Alpha + '</th>\n'
	table_head_string += '\t</tr>\n'
	table_head_string += '</thead>\n'
	table_body_string = ''
	table_body_string += '\t<tbody>\n'
	###############################################################################
	for each_Sample, each_Sample_ID in itertools.izip(Sample_name_LIST, Sample_ID_LIST):
		#FILL THE METADATA COLUMNS
		for each_SAMPLE_metadata in Sample_metadata_Columns:
			each_metadata_Value = Sample_metadata_DataFrame.loc[each_Sample_ID, each_SAMPLE_metadata]
			Alpha_Diversity_DICT[each_SAMPLE_metadata].append(each_metadata_Value)
			table_body_string += '\t\t\t<td class="text-center">' + str(each_metadata_Value) + '</td>\n'
		#Total_Number
		total_read_count = OTU_DataFrame[each_Sample].sum()
		Alpha_Diversity_DICT['Total_Number'].append(total_read_count)
		table_body_string += '\t\t\t<td class="text-center">' + str(total_read_count) + '</td>\n'
		#unclassified
		if Unclassified_Flag is True:
			unclassified_total_count_LIST = []
			unclassified_read_count = 0
			for each_INDEX in unclassified_index_LIST:
				unclassified_total_count_LIST.append(int(OTU_DataFrame.loc[each_INDEX, each_Sample]))
			unclassified_read_count = sum(unclassified_total_count_LIST)
		else:
			unclassified_read_count = 0
		#classified
		classified_read_count = int(total_read_count - unclassified_read_count)
		Alpha_Diversity_DICT['Classified_Number'].append(classified_read_count)
		table_body_string += '\t\t\t<td class="text-center">' + str(classified_read_count) + '</td>\n'
		#unclassified
		Alpha_Diversity_DICT['Unclassified_Number'].append(unclassified_read_count)
		table_body_string += '\t\t\t<td class="text-center">' + str(unclassified_read_count) + '</td>\n'
		#unclassified_percent
		unclassified_read_percentage = round_float(percentage(unclassified_read_count, total_read_count))
		Alpha_Diversity_DICT['Unclassified_Percentage'].append(unclassified_read_percentage)
		table_body_string += '\t\t\t<td class="text-center">' + str(unclassified_read_percentage) + '</td>\n'
		#classified_percent
		classified_read_percentage = round_float(percentage(classified_read_count, total_read_count))
		Alpha_Diversity_DICT['Classified_Percentage'].append(classified_read_percentage)
		table_body_string += '\t\t\t<td class="text-center">' + str(classified_read_percentage) + '</td>\n'
		#mean value
		mean_value = round_float(OTU_DataFrame[each_Sample].mean())
		Alpha_Diversity_DICT['Mean_Value'].append(mean_value)
		table_body_string += '\t\t\t<td class="text-center">' + str(mean_value) + '</td>\n'
		#Simpson_Value
		simpson_value = round_float(Alpha_DF.loc[each_Sample, 'simpson'])
		Alpha_Diversity_DICT['Simpson_diversity_index'].append(simpson_value)
		simpson_lci = round_float(Alpha_DF.loc[each_Sample, 'simpson_lci'])
		Alpha_Diversity_DICT['Simpson_Low_Confidence_Interval'].append(simpson_lci)
		simpson_hci = round_float(Alpha_DF.loc[each_Sample, 'simpson_hci'])
		Alpha_Diversity_DICT['Simpson_High_Confidence_Interval'].append(simpson_hci)
		table_body_string += '\t\t\t<td class="text-center">' + str(simpson_value) + '</td>\n'
		#Shannon_Value
		shannon_value = round_float(Alpha_DF.loc[each_Sample, 'shannon'])
		Alpha_Diversity_DICT['Shannon_diversity_index'].append(shannon_value)
		shannon_lci = round_float(Alpha_DF.loc[each_Sample, 'shannon_lci'])
		Alpha_Diversity_DICT['Shannon_Low_Confidence_Interval'].append(shannon_lci)
		shannon_hci = round_float(Alpha_DF.loc[each_Sample, 'shannon_hci'])
		Alpha_Diversity_DICT['Shannon_High_Confidence_Interval'].append(shannon_hci)
		table_body_string += '\t\t\t<td class="text-center">' + str(shannon_value) + '</td>\n'
		#max_value
		max_value = OTU_DataFrame[each_Sample].max()
		Alpha_Diversity_DICT['Maximum_Abundance_value'].append(max_value)
		table_body_string += '\t\t\t<td class="text-center">' + str(max_value) + '</td>\n'
		#min_value
		min_value = OTU_DataFrame[each_Sample].min()
		Alpha_Diversity_DICT['Minimum_Abundance_value'].append(min_value)
		table_body_string += '\t\t\t<td class="text-center">' + str(min_value) + '</td>\n'
		#max_otu
		max_otu = TAX_ID_DICT[OTU_DataFrame[each_Sample].idxmax()]
		Alpha_Diversity_DICT['Maximum_Abundance_OTU'].append(max_otu)
		table_body_string += '\t\t\t<td class="text-center">' + str(max_otu) + '</td>\n'
		#min_otu
		min_otu = TAX_ID_DICT[OTU_DataFrame[each_Sample].idxmin()]
		Alpha_Diversity_DICT['Minimum_Abundance_OTU'].append(min_otu)
		table_body_string += '\t\t\t<td class="text-center">' + str(min_otu) + '</td>\n'
		#End of html row
		table_body_string += '\t\t</tr>\n'
	table_body_string += '\t</tbody>\n'
	###############################################################################
	statistics_table_html_string = """
		<div id="ALPHA_DIVERSITY_TABLE" class="container-fluid">
			<div class="row">
				<div class="col-xs-12 col-sm-12 col-md-12 col-lg-12">
					<div class="panel panel-default">
						<div class="panel-heading" style="border:0px; background-color: #F4F9FD;">
							<h3 class="panel-title">ALPHA_DIVERSITY_TABLE</h3>
						</div>
						<div class="panel-body" align="center">
							<!-- Table -->
							<div class="table-responsive">
							<table id="statistics_table" class="table table-striped table-bordered table-hover small" cellspacing="0" width="100%" style="font-family:'Avenir';">
							""" + table_head_string + """
							""" + table_body_string + """
							</table>

						</div>
						</div>
						<div class="panel-footer">
						</div>
					</div>
				</div>
			</div>
		</div>
	"""
	FINAL_STRING = html_visualizer(FINAL_STRING, statistics_table_html_string)
	###############################################################################
	alpha_diversity_in_excel(Excel_WorkBook, Alpha_Diversity_DICT, Alpha_Diversity_LIST, 'Alpha_Diversity')
	FINAL_STRING = alpha_diversity_in_html(FINAL_STRING, Alpha_Diversity_DICT, Alpha_Diversity_LIST)
	"""
	#Create DF from the dict and write to Excel
	Alpha_DF, Alpha_DF_LIST = dict_to_pandas_dataframe_converter(Alpha_Diversity_DICT, Alpha_Diversity_LIST, the_INDEX=None)
	WorkSheet = Excel_WorkBook.create_sheet('Alpha_Diversity')
	Alpha_DF = Alpha_DF[Alpha_DF_LIST]
	for each_Row in OPEN_DF.dataframe_to_rows(Alpha_DF, index=False, header=True):
		WorkSheet.append(tuple(each_Row))
	"""
	return FINAL_STRING


def rarefaction_analysis(Excel_File_PATH, Excel_WorkBook, FINAL_STRING, mothur_exec_PATH, processors_COUNT, outputdir_PATH):
	Shared_File_PATH = outputdir_PATH + 'Shared_File.txt'
	TAX_ID_DICT = {}
	TAX_ID_DICT = excel_to_shared_file_converter(Excel_File_PATH, 'OTU', Shared_File_PATH)
	# ######################## Rarefaction curve CALCULATION
	flag, stderr = execute_functions(mothur_rarefaction_single, processors_COUNT, outputdir_PATH, 'multi', 'mothur', mothur_exec_PATH, Shared_File_PATH)
	if flag is False:
		print "Execution of mothur_rarefaction_single failed!!!"
		sys.exit(2)
	else:
		scanned_container = []
		extension_list = ['.groups.r_chao']
		flag = scandirs(outputdir_PATH, scanned_container, extension_list)
		print "Scanning.."
		if flag is False:
			print "Failed :("
			print "This extension is not available: ", extension_list
			sys.exit(2)
		else:
			print "NICE :) Found it"
			counter = 1
			for file in scanned_container:
				print "File#", str(counter), ":", file
				counter += 1
	flag = remove_extension_files(outputdir_PATH, '.rabund')
	# ######################## Rarefaction file processing
	Rarefaction_File_PATH = add_extension(scanned_container[0], '.txt')
	Rarefaction_vertical_DICT, Rarefaction_vertical_LIST = any_file_to_vertical_dict_converter(Rarefaction_File_PATH)
	X_axis_LIST = Rarefaction_vertical_DICT['numsampled']
	raw_sample_name_list = []
	for each_head in Rarefaction_vertical_LIST:
		if 'lci' in each_head:
			continue
		elif 'hci' in each_head:
			continue
		elif 'numsampled' in each_head:
			continue
		else:
			raw_sample_name_list.append(each_head)
	Rarefaction_Excel_DICT = {}
	Rarefaction_Excel_DICT['x_axis'] = map(float, X_axis_LIST)
	Rarefaction_Excel_LIST = ['x_axis']
	Rarefaction_plotly_DICT = {}
	Rarefaction_plotly_DICT['x_axis'] = map(float, X_axis_LIST)
	Rarefaction_plotly_LIST = ['x_axis']
	for each_Sample in raw_sample_name_list:
		each_Sample_LIST = each_Sample.split('-')
		clean_Sample_name = list_to_string(each_Sample_LIST[1:], '-')
		Rarefaction_Excel_DICT[clean_Sample_name] = map(float, Rarefaction_vertical_DICT[each_Sample])
		Rarefaction_plotly_DICT[clean_Sample_name] = map(float, filter_list(Rarefaction_vertical_DICT[each_Sample], 'NA'))
		Rarefaction_Excel_LIST.append(clean_Sample_name)
		Rarefaction_plotly_LIST.append(clean_Sample_name)
	# ######################
	flag = rarefaction_in_excel(Excel_WorkBook, Rarefaction_Excel_DICT, Rarefaction_Excel_LIST)
	if flag is True:
		pass
	FINAL_STRING = rarefaction_in_html(FINAL_STRING, Rarefaction_plotly_DICT, Rarefaction_plotly_LIST)
	return FINAL_STRING


def natural_abundance_analysis(Excel_File_PATH, Excel_WorkBook):
	Lineage_Name_LIST = ['Kingdom', 'Phylum', 'Class', 'Order', 'Family', 'Genus', 'Species', 'OTU']
	for each_Lineage in Lineage_Name_LIST:
		Lineage_vertical_DICT, Lineage_vertical_LIST = excel_to_vertical_dict_converter(Excel_File_PATH, each_Lineage)
		natural_abundance_in_excel(Excel_WorkBook, Lineage_vertical_DICT, Lineage_vertical_LIST, SHEETname=each_Lineage + '_abundance')
	return 1
# ################################### MAIN FUNCTION ########################## #


def main(argv):

	report_string = ''
	# ++++++++++++++++++++++++++++++ PARSE INPUT ARGUMENTS
	parser = argparse.ArgumentParser()
	main_file = parser.add_argument_group('Main file parameters')
	main_file.add_argument("--biom", help="Universal microbiome abundance matrix file(http://biom-format.org)", action='store')
	main_file.add_argument("--design", help="Design file: Tab delimited file to assign samples to a specific treatments, or other categories.", action='store')
	args = parser.parse_args()
	# ------------------------------ END OF PARSE INPUT ARGUMENTS

	# ++++++++++++++++++++++++++++++ BEURACRATICS PROCEDURES
	report_string += "######################################################################################################################################\n"
	print "######################################################################################################################################"
	report_string += "BiomeSlicer 1.0 EXECUTION HAS INITIATED" + '\n'
	print "BiomeSlicer 1.0 EXECUTION HAS INITIATED"
	report_string += "Initiation time: " + time.strftime("%Y-%m-%d %H:%M:%S") + '\n'
	print "Initiation time: ", time.strftime("%Y-%m-%d %H:%M:%S")
	report_string += "###################################################################" + '\n'
	print "###################################################################"
	report_string += "INFORMATION ABOUT THE ENVIRONMENT, EXECUTABLES AND PROVIDED DATA" + '\n'
	print "INFORMATION ABOUT THE ENVIRONMENT, EXECUTABLES AND PROVIDED DATA"
	report_string += "###################################################################" + '\n'
	print "###################################################################"
	report_string += "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++" + '\n'
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	print "COMMAND LINE:"
	report_string += "COMMAND LINE:\n"
	report_string += "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++" + '\n'
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	commandline_string = 'python ' + ' '.join(sys.argv) + '\n'
	print commandline_string
	report_string += commandline_string + '\n'
	report_string += "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++" + '\n'
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	report_string += "ARGUMENTS:" + '\n'
	print "ARGUMENTS:"
	report_string += "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++" + '\n'
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"

	# ++++++++++++++++++++++++++++++ OUTPUT DIRECTORY CHECKING
	args.outputdir = DEFAULT_OUTPUTDIR
	global report_file
	report_file = args.outputdir + "BiomeSlicer_report.txt"
	check_it_and_remove_it(report_file, True)
	report(report_string)
	# ------------------------------ END OF OUTPUT DIRECTORY CHECKING

	# ++++++++++++++++++++++++++++++ PROCESSORS CHECKING
	args.processors = DEFAULT_PROCESSORS
	# ------------------------------ END OF PROCESSORS CHECKING

	# ++++++++++++++++++++++++++++++ PREFIX NAME CHECKING
	args.prefix = DEFAULT_PREFIX
	# ------------------------------ END OF PREFIX NAME CHECKING

	# ++++++++++++++++++++++++++++++ EXECUTIVE DIRECTORY CHECKING
	args.execdir = DEFAULT_EXECDIR
	# ------------------------------ END OF EXECUTIVE DIRECTORY CHECKING

	# ++++++++++++++++++++++++++++++ CHECKING EXECUTIVES
	print "\n###################################################################"
	report("\n###################################################################")
	print "VERIFYING THE SANITY/VERSION OF EXECUTABLES IN EXECUTIVE DIRECTORY[--execdir]"
	report("VERIFYING THE SANITY/VERSION OF EXECUTABLES IN EXECUTIVE DIRECTORY[--execdir]")
	print "###################################################################\n"
	report("###################################################################\n")
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	report("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
	report("0: ENVIRONMENT")
	print "0: ENVIRONMENT"
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	report("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
	print "Operating System version is: ", platform.platform()
	report("Operating System version is: " + platform.platform())
	python_version = sys.version.split(' (')[0]
	print "Python version is: ", python_version
	report("Python version is: " + python_version)
	if float(python_version[0:3]) < 2.7:
		error("Python version is older than 2.7")
		print "Python version is older than 2.7"
		print "ABORTING!!!"
		error("ABORTING!!!")
		sys.exit(2)
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	report("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
	report("1: MOTHUR")
	print "1: MOTHUR"
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	report("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
	mothur_exec_PATH = args.execdir + 'mothur'
	if isFileExist(mothur_exec_PATH) is False:
		error("Your mothur file path has Access/Exist issue")
		print "Your mothur file path has Access/Exist issue"
		print "ABORTING!!!"
		error("ABORTING!!!")
		sys.exit(2)
	else:
		print "mothur execution file path is: ", mothur_exec_PATH
		report("mothur execution file path is: " + mothur_exec_PATH)
		#report("Testing mothur executables: ")
		print "Testing mothur executable: "
		report("Testing mothur executable: ")
		flag, stderr = execute_functions(test_mothur, args.processors, args.outputdir, 'multi', 'mothur', mothur_exec_PATH)
		if flag is False:
			error("[" + FAILED_MARK + "]")
			print "[", FAILED_MARK, "]"
			print "Execution of mothur failed!!!"
			error("Execution of mothur failed!!!")
			print "ABORTING!!!"
			error("ABORTING!!!")
			sys.exit(2)
		else:
			target_lines = []
			flag = parse_mothur_logfile(stderr, 'mothur v.', target_lines)
			if flag is False:
				print "This keyword is not avalaible: mothur v."
				error("This keyword is not avalaible: mothur v.")
				print "ABORTING!!!"
				error("ABORTING!!!")
				sys.exit(2)
			report("Mothur executables responding successfully!!!")
			print "Mothur executables responding successfully!!!"
			report("version of mothur executables: " + target_lines[0])
			print "version of mothur executables:", target_lines[0]
	print "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
	report("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
	# ------------------------------ END OF CHECKING EXECUTIVES
	# ------------------------------ END OF BEURACRATICS PROCEDURES

	# ++++++++++++++++++++++++++++++ CHECKING INPUTS
	print "\n###################################################################"
	report("\n###################################################################")
	print "VERIFYING THE SANITY/VALIDITY OF INPUT FILES"
	report("VERIFYING THE SANITY/VALIDITY OF INPUT FILES")
	print "###################################################################\n"
	report("###################################################################\n")

	# ++++++++++++++++++++++++++++++ CHECK BIOM FILE AND CONVERT TO EXCEL
	global DESIGN_EXIST
	# +++++++++++++++++++++++++++++ BIOM FILE PROCESSING
	print "BIOM FILE PROCESSING is in progress"
	report("BIOM FILE PROCESSING is in progress")
	print "Execution started at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution started at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	if args.biom is None:
		print "USING TEST BIOM FILE..."
		report("USING TEST BIOM FILE...")
		args.biom = DEFAULT_TESTDIR + "ZAC_biom.txt"
	if isFileExist(args.biom) is False:
		error("[--biom]: biom file has Access/Exist issue")
		print "[--biom]: biom file has Access/Exist issue"
		print "ABORTING!!!"
		error("ABORTING!!!")
		sys.exit(2)
	
	Excel_File_PATH = args.outputdir + 'BiomeSlicer_result.xlsx'
	flag = biom_to_excel_converter(args.biom, args.design, Excel_File_PATH)
	if flag is False:
		print "ABORTING!!!"
		error("ABORTING!!!")
	else:
		print "Execution completed at ", time.strftime("%Y-%m-%d %H:%M:%S")
		report("Execution completed at " + time.strftime("%Y-%m-%d %H:%M:%S"))
		print "BIOM FILE PROCESSING PASSED!!!"
		report("BIOM FILE PROCESSING PASSED!!!")
	# ------------------------------ END OF BIOM FILE PROCESSING

	# +++++++++++++++++++++++++++++ INITIATING
	FINAL_STRING = ''
	FINAL_STRING = html_visualizer(FINAL_STRING)
	Excel_WorkBook = openpyxl.load_workbook(Excel_File_PATH)
	# ------------------------------ END OF INITIATING

	# +++++++++++++++++++++++++++++ BRIEF STATISTICS TABLE FUNCTION
	print "STATISTICS TABLE FUNCTION is in progress"
	report("STATISTICS TABLE FUNCTION is in progress")
	print "Execution started at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution started at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	#FINAL_STRING = statistics_table_function(Excel_File_PATH, FINAL_STRING)
	#FINAL_STRING = alpha_diversity_analysis(Excel_File_PATH, FINAL_STRING)
	FINAL_STRING = alpha_diversity_analysis(Excel_File_PATH, Excel_WorkBook, FINAL_STRING, mothur_exec_PATH, args.processors, args.outputdir)
	print "STATISTICS TABLE FUNCTION PASSED!!!"
	report("STATISTICS TABLE FUNCTION PASSED!!!")
	print "Execution completed at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution completed at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	# ----------------------------- END OF BRIEF STATISTICS TABLE
	"""
	
	# +++++++++++++++++++++++++++++ RAREFACTION CURVE LINECHART FUNCTION
	print "RAREFACTION CURVE LINECHART FUNCTION is in progress"
	report("RAREFACTION CURVE LINECHART FUNCTION is in progress")
	print "Execution started at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution started at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	FINAL_STRING = rarefaction_analysis(Excel_File_PATH, Excel_WorkBook, FINAL_STRING, mothur_exec_PATH, args.processors, args.outputdir)
	print "RAREFACTION CURVE LINECHART FUNCTION PASSED!!!"
	report("RAREFACTION CURVE LINECHART FUNCTION PASSED!!!")
	print "Execution completed at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution completed at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	# ----------------------------- END OF RAREFACTION CURVE LINECHART FUNCTION

	# +++++++++++++++++++++++++++++ NATURAL ABUNDANCE BARPLOT FUNCTION
	print "NATURAL ABUNDANCE BARPLOT FUNCTION is in progress"
	report("NATURAL ABUNDANCE BARPLOT FUNCTION is in progress")
	print "Execution started at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution started at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	natural_abundance_analysis(Excel_File_PATH, Excel_WorkBook)
	print "NATURAL ABUNDANCE BARPLOT FUNCTION PASSED!!!"
	report("NATURAL ABUNDANCE BARPLOT FUNCTION PASSED!!!")
	print "Execution completed at ", time.strftime("%Y-%m-%d %H:%M:%S")
	report("Execution completed at " + time.strftime("%Y-%m-%d %H:%M:%S"))
	# ----------------------------- END OF NATURAL ABUNDANCE BARPLOT FUNCTION
	
	
	"""
	# +++++++++++++++++++++++++++++SAVING WORKBOOK
	Excel_WorkBook.save(Excel_File_PATH)
	# -----------------------------END OF SAVING WORKBOOK
	# +++++++++++++++++++++++++++++ FINALIZING
	FINAL_STRING = html_visualizer(FINAL_STRING)
	flag = remove_extension_files(CURRENT_PATH, '.logfile')
	write_string_down(FINAL_STRING, 'BiomeSlicer_result.html')
	zip_file_NAME = 'BiomeSlicer_result.zip'
	zip_it('BiomeSlicer_result.html', zip_file_NAME)
	print "BIOME SLICER EXECUTION COMPLETED."
	report("BIOME SLICER EXECUTION COMPLETED.")
	report("Completion time: " + time.strftime("%Y-%m-%d %H:%M:%S"))
	print "Completion time: ", time.strftime("%Y-%m-%d %H:%M:%S")
	# ----------------------------- END OF FINALIZING


# ################################### FINITO ################################# #
if __name__ == "__main__": main(sys.argv[1:])
